# -*- coding: utf-8 -*-
"""
Genera un Excel de demo mostrando el formato exacto del informe OGSA Paso 5.
Ejecutar: python generate_demo_excel.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random

# ── Paleta OAG ────────────────────────────────────────────────────────────────
COLOR_HEADER      = "1C2B3A"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_SYNGENTA    = "1A4A8A"
COLOR_ZEBRA       = "F0F4F8"
COLOR_SYNGENTA_ROW = "EBF0F9"
COLOR_DIF         = "FFF3CD"
COLOR_SOLO        = "FDECEA"
COLOR_OK_ROW      = "EAFAF1"
COLOR_TOTAL       = "E8EDF3"
COLOR_BORDER      = "DDE1E7"

MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
         "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

# ── Datos de demo ─────────────────────────────────────────────────────────────
EXPEDIENTE = {
    "nombre_distribuidor": "AGRO DEL SUR S.R.L.",
    "cuit_distribuidor":   "30-71234567-8",
    "anio_analisis":       2024,
    "auditor":             "Lic. María González",
}

# Compras por proveedor (USD)
COMPRAS = [
    {"Proveedor": "SYNGENTA ARGENTINA S.A.",        "CUIT": "30-61234500-1",
     "E":148320, "F":162450, "M":201340, "A":189760, "My":178920, "Jn":165430,
     "Jl":143210, "Ag":156780, "S":198340, "O":212450, "N":187630, "D":94120},
    {"Proveedor": "BAYER S.A.",                     "CUIT": "30-52345601-2",
     "E":87230,  "F":92140,  "M":118760, "A":103420, "My":98340,  "Jn":89210,
     "Jl":78430, "Ag":84320, "S":109870, "O":121340, "N":98760,   "D":48320},
    {"Proveedor": "BASF ARGENTINA S.A.",            "CUIT": "30-63456702-3",
     "E":54120,  "F":61340,  "M":78920,  "A":72310,  "My":68450,  "Jn":58930,
     "Jl":49870, "Ag":53210, "S":71240,  "O":82340,  "N":67890,   "D":31240},
    {"Proveedor": "NUFARM ARGENTINA S.A.",          "CUIT": "30-74567803-4",
     "E":38910,  "F":42340,  "M":56780,  "A":51230,  "My":47890,  "Jn":41230,
     "Jl":34560, "Ag":37890, "S":49870,  "O":58340,  "N":46780,   "D":22140},
    {"Proveedor": "FMC ARGENTINA S.A.",             "CUIT": "30-85678904-5",
     "E":28340,  "F":31230,  "M":41560,  "A":38920,  "My":35670,  "Jn":30120,
     "Jl":25430, "Ag":27890, "S":36780,  "O":42340,  "N":34120,   "D":16780},
    {"Proveedor": "YPF AGRO S.A.",                  "CUIT": "30-96789005-6",
     "E":19230,  "F":21340,  "M":28760,  "A":26540,  "My":24310,  "Jn":20870,
     "Jl":17560, "Ag":19230, "S":25430,  "O":29870,  "N":23450,   "D":11230},
    {"Proveedor": "DOW AGROSCIENCES ARG. S.A.",     "CUIT": "30-07890106-7",
     "E":14230,  "F":15670,  "M":20340,  "A":18760,  "My":17230,  "Jn":14890,
     "Jl":12450, "Ag":13780, "S":18340,  "O":21230,  "N":16780,   "D":8120},
    {"Proveedor": "RIZOBACTER ARGENTINA S.A.",      "CUIT": "30-18901207-8",
     "E":8920,   "F":9870,   "M":12780,  "A":11560,  "My":10890,  "Jn":9340,
     "Jl":7890,  "Ag":8560,  "S":11230,  "O":13450,  "N":10560,   "D":4780},
]

for row in COMPRAS:
    vals = [row["E"],row["F"],row["M"],row["A"],row["My"],row["Jn"],
            row["Jl"],row["Ag"],row["S"],row["O"],row["N"],row["D"]]
    for mes, val in zip(MESES, vals):
        row[mes] = val
    row["Total"] = sum(vals)

# Venta agroquímicos (Anexo II)
APERTURA = [
    {"articulo":"ACTARA 25 WG x 1 KG",     "syngenta":"SI",
     "E":18420,"F":21340,"M":28760,"A":24890,"My":19870,"Jn":17340,
     "Jl":15230,"Ag":17890,"S":23450,"O":27890,"N":21230,"D":9870},
    {"articulo":"AMISTAR TOP 325 SC x 5 L", "syngenta":"SI",
     "E":24310,"F":28450,"M":36780,"A":32340,"My":27890,"Jn":23450,
     "Jl":19870,"Ag":22340,"S":31230,"O":37890,"N":29870,"D":13450},
    {"articulo":"ENGEO 247 SC x 1 L",       "syngenta":"SI",
     "E":12340,"F":14560,"M":19870,"A":17230,"My":15670,"Jn":12890,
     "Jl":10780,"Ag":12340,"S":16780,"O":20340,"N":15670,"D":7230},
    {"articulo":"KARATE ZEON 250 CS x 1 L", "syngenta":"SI",
     "E":9870, "F":11230,"M":15670,"A":13890,"My":12340,"Jn":10230,
     "Jl":8560, "Ag":9870, "S":13230,"O":16780,"N":12890,"D":5670},
    {"articulo":"ELATUS ERA x 5 L",         "syngenta":"SI",
     "E":8230, "F":9450, "M":12780,"A":11230,"My":10120,"Jn":8560,
     "Jl":7120, "Ag":8230, "S":10890,"O":13450,"N":10230,"D":4780},
    {"articulo":"TROPHY 500 EC x 20 L",     "syngenta":"SI",
     "E":6780, "F":7890, "M":10340,"A":9120, "My":8230, "Jn":6890,
     "Jl":5780, "Ag":6780, "S":8890, "O":10780,"N":8340, "D":3890},
    {"articulo":"ROUNDUP ULTRA MAX x 20 L", "syngenta":"NO",
     "E":31230,"F":36780,"M":48340,"A":42890,"My":38230,"Jn":32340,
     "Jl":27890,"Ag":31230,"S":42340,"O":51230,"N":39870,"D":18450},
    {"articulo":"2,4 D ADVANCE 48 SL x 20L","syngenta":"NO",
     "E":14560,"F":16780,"M":22340,"A":19870,"My":17890,"Jn":14780,
     "Jl":12340,"Ag":14560,"S":19450,"O":23450,"N":18230,"D":8560},
    {"articulo":"WARRANT 350 CS x 5 L",     "syngenta":"NO",
     "E":8340, "F":9560, "M":12890,"A":11230,"My":10120,"Jn":8560,
     "Jl":7120, "Ag":8340, "S":11230,"O":13230,"N":10230,"D":4780},
    {"articulo":"CIPERMETRINA 25% EC x 1L", "syngenta":"NO",
     "E":5670, "F":6780, "M":8920, "A":7890, "My":7120, "Jn":5890,
     "Jl":4890, "Ag":5670, "S":7560, "O":9230, "N":7120, "D":3340},
]
for row in APERTURA:
    vals = [row["E"],row["F"],row["M"],row["A"],row["My"],row["Jn"],
            row["Jl"],row["Ag"],row["S"],row["O"],row["N"],row["D"]]
    for mes, val in zip(MESES, vals):
        row[mes] = val
    row["Total"] = sum(vals)

# Cruce CRM (Anexo III)
CRM = [
    {"producto":"ACTARA 25 WG x 1 KG",     "fecha":"15/03/2024","tipo":"FAC","comprobante":"0001-00012456",
     "cuit":"20-12345678-9","cant_g":250.0,"cant_crm":250.0,"delta_cant":0.0,
     "monto_g":28760.0,"monto_crm":28760.0,"delta":0.0,
     "estado":"OK","justificacion":"Registros coinciden en cantidad y monto."},
    {"producto":"AMISTAR TOP 325 SC x 5 L","fecha":"22/03/2024","tipo":"FAC","comprobante":"0001-00012489",
     "cuit":"20-23456789-0","cant_g":180.0,"cant_crm":175.0,"delta_cant":-5.0,
     "monto_g":36780.0,"monto_crm":35760.0,"delta":-1020.0,
     "estado":"DIFERENCIA","justificacion":"Diferencia de 5 unidades por devolución parcial no registrada en CRM al cierre del período. Ajuste pendiente de confirmación con el distribuidor."},
    {"producto":"ENGEO 247 SC x 1 L",      "fecha":"08/04/2024","tipo":"FAC","comprobante":"0001-00012534",
     "cuit":"20-34567890-1","cant_g":420.0,"cant_crm":420.0,"delta_cant":0.0,
     "monto_g":19870.0,"monto_crm":19870.0,"delta":0.0,
     "estado":"OK","justificacion":"Registros coinciden en cantidad y monto."},
    {"producto":"KARATE ZEON 250 CS x 1 L","fecha":"18/04/2024","tipo":"FAC","comprobante":"0001-00012578",
     "cuit":"20-45678901-2","cant_g":310.0,"cant_crm":None,"delta_cant":None,
     "monto_g":15670.0,"monto_crm":None,"delta":None,
     "estado":"SOLO_GESTION","justificacion":"Comprobante no registrado en CRM. Posible omisión de carga por parte del representante de ventas. Se recomienda verificar con el área comercial de Syngenta."},
    {"producto":"ELATUS ERA x 5 L",        "fecha":"05/05/2024","tipo":"FAC","comprobante":"0001-00012623",
     "cuit":"20-56789012-3","cant_g":145.0,"cant_crm":145.0,"delta_cant":0.0,
     "monto_g":12780.0,"monto_crm":12780.0,"delta":0.0,
     "estado":"OK","justificacion":"Registros coinciden en cantidad y monto."},
    {"producto":"TROPHY 500 EC x 20 L",    "fecha":"14/05/2024","tipo":"FAC","comprobante":"0001-00012667",
     "cuit":"20-67890123-4","cant_g":89.0,"cant_crm":95.0,"delta_cant":6.0,
     "monto_g":10340.0,"monto_crm":11038.0,"delta":698.0,
     "estado":"DIFERENCIA","justificacion":"CRM registra 6 unidades adicionales correspondientes a una entrega complementaria de la misma operación, facturada por separado en nota de débito 0001-00002341."},
    {"producto":"ACTARA 25 WG x 1 KG",     "fecha":"20/06/2024","tipo":"FAC","comprobante":"0001-00012789",
     "cuit":"20-78901234-5","cant_g":None,"cant_crm":200.0,"delta_cant":None,
     "monto_g":None,"monto_crm":23120.0,"delta":None,
     "estado":"SOLO_CRM","justificacion":"Operación registrada en CRM sin comprobante en gestión. El distribuidor informa que corresponde a una venta de muestra sin cargo facturada directamente por Syngenta Argentina. No impacta en los estados contables del DS."},
    {"producto":"AMISTAR TOP 325 SC x 5 L","fecha":"12/07/2024","tipo":"NC","comprobante":"0001-00002123",
     "cuit":"20-23456789-0","cant_g":-30.0,"cant_crm":-30.0,"delta_cant":0.0,
     "monto_g":-6140.0,"monto_crm":-6140.0,"delta":0.0,
     "estado":"OK","justificacion":"Nota de crédito por devolución. Registros coinciden."},
    {"producto":"ELATUS ERA x 5 L",        "fecha":"03/09/2024","tipo":"FAC","comprobante":"0001-00013012",
     "cuit":"20-89012345-6","cant_g":230.0,"cant_crm":230.0,"delta_cant":0.0,
     "monto_g":21340.0,"monto_crm":21340.0,"delta":0.0,
     "estado":"OK","justificacion":"Registros coinciden en cantidad y monto."},
    {"producto":"KARATE ZEON 250 CS x 1 L","fecha":"25/10/2024","tipo":"FAC","comprobante":"0001-00013298",
     "cuit":"20-90123456-7","cant_g":178.0,"cant_crm":165.0,"delta_cant":-13.0,
     "monto_g":16780.0,"monto_crm":15555.0,"delta":-1225.0,
     "estado":"DIFERENCIA","justificacion":"Diferencia atribuible a descuento comercial especial aplicado por Syngenta en CRM no reflejado en el comprobante fiscal. Documentación de respaldo adjunta en carpeta de auditoría."},
]

# ── Helpers de estilo ─────────────────────────────────────────────────────────
def thin_border(color=COLOR_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def header_style(color=COLOR_HEADER, font_color=COLOR_HEADER_FONT):
    return {
        "fill": PatternFill("solid", fgColor=color),
        "font": Font(name="Calibri", bold=True, color=font_color, size=10),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": thin_border("FFFFFF"),
    }

def apply(cell, **kwargs):
    for k, v in kwargs.items():
        setattr(cell, k, v)

def write_header(ws, row, cols, color=COLOR_HEADER, font_color=COLOR_HEADER_FONT):
    st = header_style(color, font_color)
    for ci, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=val)
        apply(c, **st)
    ws.row_dimensions[row].height = 28

def data_cell(ws, row, col, val, zebra=False, fmt=None, bold=False,
              color=None, align=None, fc=None, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    bg = color or (COLOR_ZEBRA if zebra else "FFFFFF")
    c.fill = PatternFill("solid", fgColor=bg)
    fc_ = fc or ("000000" if not bold else "1C2B3A")
    c.font = Font(name="Calibri", size=9, bold=bold, color=fc_)
    h_align = align or ("right" if isinstance(val, (int, float)) and val is not None else "left")
    c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
    c.border = thin_border("E5E9EF")
    if fmt: c.number_format = fmt
    ws.row_dimensions[row].height = 15 if not wrap else 40
    return c

def titulo_hoja(ws, titulo, sub=""):
    ws.sheet_view.showGridLines = False
    # Banda superior decorativa
    for col in range(1, 30):
        ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor=COLOR_HEADER)
        ws.cell(row=2, column=col).fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 6

    ws.merge_cells("A3:Z3")
    c = ws["A3"]
    c.value = titulo
    c.font = Font(name="Calibri", bold=True, size=13, color=COLOR_HEADER)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 22

    ws.merge_cells("A4:Z4")
    s = ws["A4"]
    s.value = sub
    s.font = Font(name="Calibri", size=9, color="777777")
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 14

    # Línea separadora
    for col in range(1, 30):
        ws.cell(row=5, column=col).fill = PatternFill("solid", fgColor=COLOR_SYNGENTA)
    ws.row_dimensions[5].height = 2

NUM_FMT = '#,##0.00'
PCT_FMT = '0.0%'

# ── PORTADA ───────────────────────────────────────────────────────────────────
def hoja_portada(wb):
    ws = wb.create_sheet("Portada")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 50

    # Header
    for r in range(1, 10):
        for col in range(1, 20):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=COLOR_HEADER)
        ws.row_dimensions[r].height = 16 if r not in (3, 4) else 24

    # Línea dorada
    for col in range(1, 20):
        ws.cell(row=9, column=col).fill = PatternFill("solid", fgColor="C8A84B")
    ws.row_dimensions[9].height = 3

    ws.merge_cells("B3:K4")
    t = ws["B3"]
    t.value = "OGSA — SISTEMA DE AUDITORÍAS COMERCIALES"
    t.font = Font(name="Calibri", bold=True, size=17, color="FFFFFF")
    t.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B5:K6")
    s = ws["B5"]
    s.value = "Informe Ejecutivo de Compras y Ventas — Paso 5"
    s.font = Font(name="Calibri", size=11, color="9CB8D8")
    s.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B7:K8")
    tag = ws["B7"]
    tag.value = "Documento generado automáticamente por el Sistema de Auditorías OGSA"
    tag.font = Font(name="Calibri", size=8, italic=True, color="6B8BA8")
    tag.alignment = Alignment(horizontal="left", vertical="center")

    # Datos
    datos = [
        ("Distribuidor auditado:", EXPEDIENTE["nombre_distribuidor"]),
        ("CUIT:",                  EXPEDIENTE["cuit_distribuidor"]),
        ("Período analizado:",     str(EXPEDIENTE["anio_analisis"])),
        ("Auditor responsable:",   EXPEDIENTE["auditor"]),
        ("Fecha de emisión:",      "22/05/2026"),
        ("", ""),
        ("Contenido:", ""),
        ("  Anexo I —",  "Resumen de Compras por Proveedor (Top 90% del total en USD)"),
        ("  Anexo II —", "Venta de Agroquímicos — Tabla de Apertura Mensual (en USD)"),
        ("  Anexo III —","Cruce de Información CRM Syngenta vs. Gestión"),
    ]

    for i, (label, val) in enumerate(datos, start=12):
        ws.row_dimensions[i].height = 16
        if label.startswith("  Anexo"):
            cl = ws.cell(row=i, column=2, value=label)
            cl.font = Font(name="Calibri", bold=True, size=9, color=COLOR_SYNGENTA)
            cv = ws.cell(row=i, column=3, value=val)
            cv.font = Font(name="Calibri", size=9)
        elif label == "Contenido:":
            cl = ws.cell(row=i, column=2, value=label)
            cl.font = Font(name="Calibri", bold=True, size=10, color=COLOR_HEADER)
        elif label == "":
            pass
        else:
            cl = ws.cell(row=i, column=2, value=label)
            cl.font = Font(name="Calibri", bold=True, size=10)
            cv = ws.cell(row=i, column=3, value=val)
            cv.font = Font(name="Calibri", size=10)

    # Nota al pie
    ws.row_dimensions[25].height = 14
    n = ws.cell(row=25, column=2,
                value="Documento confidencial — uso exclusivo de OGSA Auditores y el cliente auditado.")
    n.font = Font(name="Calibri", size=8, italic=True, color="999999")


# ── ANEXO I — COMPRAS ─────────────────────────────────────────────────────────
def hoja_compras(wb):
    ws = wb.create_sheet("Anexo I — Compras")
    titulo_hoja(ws,
                "ANEXO I — RESUMEN DE COMPRAS POR PROVEEDOR",
                f"Compras correspondientes al período {EXPEDIENTE['anio_analisis']} · "
                f"Distribuidor: {EXPEDIENTE['nombre_distribuidor']} · "
                "Montos en USD · Incluye proveedores que representan el 90% del total")

    # Headers
    hdrs = ["Proveedor", "CUIT"] + MESES + ["TOTAL USD", "% del Total"]
    write_header(ws, 6, hdrs)

    total_general = sum(r["Total"] for r in COMPRAS)

    for i, row in enumerate(COMPRAS):
        r = i + 7
        zebra = i % 2 == 0
        z = COLOR_ZEBRA if zebra else "FFFFFF"

        data_cell(ws, r, 1, row["Proveedor"], color=z, bold=True)
        data_cell(ws, r, 2, row["CUIT"], color=z, align="center")
        for m_idx, mes in enumerate(MESES):
            v = row[mes]
            data_cell(ws, r, 3 + m_idx, v, color=z, fmt=NUM_FMT)
        data_cell(ws, r, 15, row["Total"], color=z, fmt=NUM_FMT, bold=True)
        pct = row["Total"] / total_general if total_general else 0
        data_cell(ws, r, 16, pct, color=z, fmt=PCT_FMT)

    # Fila total
    tr = len(COMPRAS) + 7
    ws.row_dimensions[tr].height = 18
    c = data_cell(ws, tr, 1, "TOTAL GENERAL", color=COLOR_TOTAL, bold=True)
    data_cell(ws, tr, 2, "", color=COLOR_TOTAL)
    for m_idx, mes in enumerate(MESES):
        s = sum(r[mes] for r in COMPRAS)
        data_cell(ws, tr, 3 + m_idx, s, color=COLOR_TOTAL, fmt=NUM_FMT, bold=True)
    data_cell(ws, tr, 15, total_general, color=COLOR_TOTAL, fmt=NUM_FMT, bold=True)
    data_cell(ws, tr, 16, 1.0, color=COLOR_TOTAL, fmt=PCT_FMT, bold=True)

    # Anchos
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    for col in range(3, 17):
        ws.column_dimensions[get_column_letter(col)].width = 13

    # Freeze
    ws.freeze_panes = "C7"


# ── ANEXO II — AGROQUÍMICOS ───────────────────────────────────────────────────
def hoja_apertura(wb):
    ws = wb.create_sheet("Anexo II — Agroquímicos")
    titulo_hoja(ws,
                "ANEXO II — VENTA DE AGROQUÍMICOS — TABLA DE APERTURA",
                f"Facturación neta en USD por producto y mes · Período {EXPEDIENTE['anio_analisis']} · "
                f"Distribuidor: {EXPEDIENTE['nombre_distribuidor']}")

    hdrs = ["Artículo / Producto", "Syngenta"] + MESES + ["TOTAL USD"]
    write_header(ws, 6, hdrs)

    # Leyenda
    ws.merge_cells("A5:O5")
    leg = ws["A5"]
    leg.value = (
        "   Productos Syngenta (azul claro)   |   "
        "Otros fabricantes (blanco/gris)   |   "
        "Nota: montos en USD, negativos = notas de crédito"
    )
    leg.font = Font(name="Calibri", size=8, italic=True, color="555555")
    leg.fill = PatternFill("solid", fgColor="F0F6FF")
    ws.row_dimensions[5].height = 13

    total_syn = 0
    total_otros = 0

    for i, row in enumerate(APERTURA):
        r = i + 7
        es_syn = row["syngenta"] == "SI"
        bg = COLOR_SYNGENTA_ROW if es_syn else (COLOR_ZEBRA if i % 2 == 0 else "FFFFFF")

        c1 = data_cell(ws, r, 1, row["articulo"], color=bg, bold=es_syn)
        c2 = data_cell(ws, r, 2, row["syngenta"], color=bg, align="center",
                       bold=es_syn, fc=COLOR_SYNGENTA if es_syn else "666666")

        for m_idx, mes in enumerate(MESES):
            v = row[mes]
            data_cell(ws, r, 3 + m_idx, v, color=bg, fmt=NUM_FMT)

        c_tot = data_cell(ws, r, 15, row["Total"], color=bg, fmt=NUM_FMT, bold=True)
        if es_syn:
            c_tot.font = Font(name="Calibri", size=9, bold=True, color=COLOR_SYNGENTA)
            total_syn += row["Total"]
        else:
            total_otros += row["Total"]

    # Subtotales
    r_syn = len(APERTURA) + 7
    ws.row_dimensions[r_syn].height = 16
    data_cell(ws, r_syn, 1, "SUBTOTAL SYNGENTA", color="D6E4F7", bold=True, fc=COLOR_SYNGENTA)
    data_cell(ws, r_syn, 2, "SI", color="D6E4F7", bold=True, fc=COLOR_SYNGENTA, align="center")
    for m_idx, mes in enumerate(MESES):
        s = sum(r[mes] for r in APERTURA if r["syngenta"] == "SI")
        data_cell(ws, r_syn, 3 + m_idx, s, color="D6E4F7", fmt=NUM_FMT, bold=True, fc=COLOR_SYNGENTA)
    data_cell(ws, r_syn, 15, total_syn, color="D6E4F7", fmt=NUM_FMT, bold=True, fc=COLOR_SYNGENTA)

    r_otros = r_syn + 1
    ws.row_dimensions[r_otros].height = 16
    data_cell(ws, r_otros, 1, "SUBTOTAL OTROS FABRICANTES", color=COLOR_ZEBRA, bold=True)
    data_cell(ws, r_otros, 2, "NO", color=COLOR_ZEBRA, align="center")
    for m_idx, mes in enumerate(MESES):
        s = sum(r[mes] for r in APERTURA if r["syngenta"] == "NO")
        data_cell(ws, r_otros, 3 + m_idx, s, color=COLOR_ZEBRA, fmt=NUM_FMT, bold=True)
    data_cell(ws, r_otros, 15, total_otros, color=COLOR_ZEBRA, fmt=NUM_FMT, bold=True)

    r_total = r_otros + 1
    ws.row_dimensions[r_total].height = 18
    data_cell(ws, r_total, 1, "TOTAL GENERAL AGROQUÍMICOS", color=COLOR_TOTAL, bold=True)
    data_cell(ws, r_total, 2, "", color=COLOR_TOTAL)
    for m_idx, mes in enumerate(MESES):
        s = sum(r[mes] for r in APERTURA)
        data_cell(ws, r_total, 3 + m_idx, s, color=COLOR_TOTAL, fmt=NUM_FMT, bold=True)
    data_cell(ws, r_total, 15, total_syn + total_otros, color=COLOR_TOTAL, fmt=NUM_FMT, bold=True)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 10
    for col in range(3, 16):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.freeze_panes = "C7"


# ── ANEXO III — CRM ───────────────────────────────────────────────────────────
def hoja_crm(wb):
    ws = wb.create_sheet("Anexo III — CRM")
    titulo_hoja(ws,
                "ANEXO III — CRUCE CRM SYNGENTA vs. GESTIÓN",
                f"Comparación de ventas de agroquímicos Syngenta · Período {EXPEDIENTE['anio_analisis']} · "
                f"Distribuidor: {EXPEDIENTE['nombre_distribuidor']}")

    # Leyenda estados
    ws.merge_cells("A5:L5")
    leg = ws["A5"]
    leg.value = (
        "   ✔ OK: coincide gestión y CRM   |   "
        "⚠ DIFERENCIA: montos o cantidades distintos (amarillo)   |   "
        "✗ SOLO GESTIÓN / SOLO CRM: registro en uno solo de los sistemas (rojo claro)"
    )
    leg.font = Font(name="Calibri", size=8, italic=True, color="444444")
    leg.fill = PatternFill("solid", fgColor="FFFBF0")
    ws.row_dimensions[5].height = 13

    hdrs = [
        "Producto", "Fecha", "Tipo", "N° Comprobante", "CUIT Cliente",
        "Cant. Gestión", "Cant. CRM", "Δ Cantidad",
        "Monto Gestión USD", "Monto CRM USD", "Δ Monto USD",
        "Justificación / Observaciones",
    ]
    write_header(ws, 6, hdrs, color=COLOR_SYNGENTA)

    ok_count = dif_count = solo_count = 0

    for i, row in enumerate(CRM):
        r = i + 7
        estado = row["estado"]

        if estado == "OK":
            bg = "EAF7F0" if i % 2 == 0 else "F5FBF8"
            ok_count += 1
        elif estado == "DIFERENCIA":
            bg = COLOR_DIF
            dif_count += 1
        else:
            bg = COLOR_SOLO
            solo_count += 1

        data_cell(ws, r, 1,  row["producto"],      color=bg, bold=(estado != "OK"))
        data_cell(ws, r, 2,  row["fecha"],          color=bg, align="center")
        data_cell(ws, r, 3,  row["tipo"],           color=bg, align="center")
        data_cell(ws, r, 4,  row["comprobante"],    color=bg)
        data_cell(ws, r, 5,  row["cuit"],           color=bg)
        data_cell(ws, r, 6,  row["cant_g"],         color=bg, fmt='#,##0.0')
        data_cell(ws, r, 7,  row["cant_crm"],       color=bg, fmt='#,##0.0')
        d_cant = row["delta_cant"]
        c8 = data_cell(ws, r, 8, d_cant, color=bg, fmt='#,##0.0',
                       fc=("C0392B" if d_cant and d_cant != 0 else None))
        data_cell(ws, r, 9,  row["monto_g"],        color=bg, fmt=NUM_FMT)
        data_cell(ws, r, 10, row["monto_crm"],      color=bg, fmt=NUM_FMT)
        d_monto = row["delta"]
        c11 = data_cell(ws, r, 11, d_monto, color=bg, fmt=NUM_FMT,
                        bold=(estado == "DIFERENCIA"),
                        fc=("C0392B" if d_monto and d_monto != 0 else None))
        data_cell(ws, r, 12, row["justificacion"],  color=bg, wrap=True)
        ws.row_dimensions[r].height = 45

    # Resumen
    r_res = len(CRM) + 8
    ws.row_dimensions[r_res].height = 14
    ws.merge_cells(f"A{r_res}:L{r_res}")
    s_cell = ws[f"A{r_res}"]
    s_cell.value = (f"Resumen: {ok_count} OK  ·  {dif_count} con diferencia  ·  "
                    f"{solo_count} solo en un sistema  ·  Total {len(CRM)} registros analizados")
    s_cell.font = Font(name="Calibri", bold=True, size=9, color=COLOR_HEADER)
    s_cell.fill = PatternFill("solid", fgColor=COLOR_TOTAL)
    s_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 17
    for col in ["F","G","H"]:
        ws.column_dimensions[col].width = 13
    for col in ["I","J","K"]:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["L"].width = 52
    ws.freeze_panes = "A7"


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    wb.remove(wb.active)

    hoja_portada(wb)
    hoja_compras(wb)
    hoja_apertura(wb)
    hoja_crm(wb)

    output = "OGSA_Informe_Demo.xlsx"
    wb.save(output)
    print(f"Excel generado: {output}")

if __name__ == "__main__":
    main()
