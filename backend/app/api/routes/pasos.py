"""
Rutas unificadas para ejecutar los 6 pasos de auditoría.
"""
import os
import io
import json
import zipfile
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from typing import List, Optional

from ...core.database import get_db
from ...core.deps import get_current_user
from ...models.user import User
from ...models.expediente import (
    Expediente, Archivo, ResultadoPaso, MaestroSyngenta, Glosario,
    AnotacionConciliacion, TipoArchivo, EstadoExpediente
)
from ...services import paso1_service, paso2_service, paso3_service
from ...services import paso4_service, paso5_service, paso6_service
from ...ai.validator import validar_paso, combinar_validacion
from ...core.logger import log as _log
import time

router = APIRouter(prefix="/expedientes/{exp_id}/pasos", tags=["pasos"])


def _get_exp(exp_id, db, user):
    exp = db.query(Expediente).filter(Expediente.id == exp_id).first()
    if not exp:
        raise HTTPException(404, "Expediente no encontrado")
    if exp.user_id != user.id and user.role.value != "ADMIN":
        raise HTTPException(403, "Sin acceso")
    return exp


def _get_archivo_path(exp_id, tipo, db) -> str:
    a = db.query(Archivo).filter(
        Archivo.expediente_id == exp_id,
        Archivo.tipo == tipo,
    ).first()
    if not a:
        raise HTTPException(400, f"Archivo {tipo.value} no cargado")
    if not os.path.exists(a.path):
        raise HTTPException(
            400,
            f"El archivo {tipo.value} se perdió del servidor (probablemente por un "
            f"redeploy). Volvé a subirlo desde el expediente."
        )
    return a.path


def _save_resultado(db, exp_id, paso, subtipo, datos=None, archivo_path=None):
    existing = db.query(ResultadoPaso).filter(
        ResultadoPaso.expediente_id == exp_id,
        ResultadoPaso.paso == paso,
        ResultadoPaso.subtipo == subtipo,
    ).first()
    if existing:
        existing.datos = datos
        existing.archivo_path = archivo_path
    else:
        db.add(ResultadoPaso(
            expediente_id=exp_id,
            paso=paso,
            subtipo=subtipo,
            datos=datos,
            archivo_path=archivo_path,
        ))
    db.commit()


def _marcar_paso_completado(exp, paso, db):
    # IMPORTANTE: hacer una COPIA de la lista. Si modificamos in-place la lista
    # existente y reasignamos la misma referencia, SQLAlchemy no detecta el
    # cambio en la columna JSON y el commit es un no-op. Esto hacía que solo
    # Paso 1 se marcara (porque arrancaba en None) y los siguientes nunca.
    completados = list(exp.pasos_completados or [])
    if paso not in completados:
        completados.append(paso)
    exp.pasos_completados = completados
    exp.paso_actual = max(completados) + 1 if completados else 1
    if 6 in completados:
        exp.estado = EstadoExpediente.COMPLETADO
    elif completados:
        exp.estado = EstadoExpediente.EN_PROCESO
    db.commit()


def _log_paso_inicio(db, exp_id, user_id, paso):
    _log(db, expediente_id=exp_id, user_id=user_id, paso=paso,
         nivel="info", evento="paso_iniciado", mensaje=f"Paso {paso} iniciado")


def _log_paso_fin(db, exp_id, user_id, paso, t0, contexto=None):
    _log(db, expediente_id=exp_id, user_id=user_id, paso=paso,
         nivel="info", evento="paso_completado",
         mensaje=f"Paso {paso} ejecutado correctamente",
         contexto=contexto, duracion_ms=int((time.time() - t0) * 1000))


def _log_paso_error(db, exp_id, user_id, paso, t0, exc, tb_str=None):
    _log(db, expediente_id=exp_id, user_id=user_id, paso=paso,
         nivel="error", evento="paso_error",
         mensaje=f"Paso {paso} falló: {type(exc).__name__}: {str(exc)[:500] or '(sin mensaje)'}",
         contexto={"traceback": tb_str} if tb_str else None,
         duracion_ms=int((time.time() - t0) * 1000))


def _safe_validar(db, exp_id: int, paso: int, build_validacion):
    """
    Corre la validación (IA + guardas) de forma defensiva. Si algo falla
    (red, API, serialización, etc.) NO bloquea el avance del paso — devuelve
    una validación vacía y deja un mensaje en logs.

    Llamarse SIEMPRE después de _marcar_paso_completado para que el paso quede
    marcado como hecho aunque la validación falle.

    Agrega automáticamente cross-step consistency checks (Capa 2).
    """
    try:
        validacion = build_validacion() or {}
        # ── Cross-step checks ──────────────────────────────────────────
        try:
            extras = _cross_step_checks(db, exp_id, paso)
            if extras:
                validacion = dict(validacion or {})
                alertas = list(validacion.get("alertas") or [])
                alertas.extend(extras)
                validacion["alertas"] = alertas
                # Si hay alguna alerta crítica, elevar la severidad global
                if any(a.get("nivel") == "error" for a in extras):
                    validacion["severidad"] = "error"
                elif any(a.get("nivel") == "warning" for a in extras) \
                        and validacion.get("severidad") in (None, "info"):
                    validacion["severidad"] = "warning"
        except Exception as e:
            print(f"[PASO {paso}] cross-step checks fallaron (no bloquea): {e}")

        _save_resultado(db, exp_id, paso, "validacion", datos=validacion)
        return validacion
    except Exception as e:
        print(f"[PASO {paso}] Validación falló — el paso queda igualmente completado. Error: {e}")
        return {"ok": True, "severidad": "info", "alertas": []}


def _get_resumen_paso(db: Session, exp_id: int, paso: int) -> dict:
    """Lee el subtipo 'resumen' o 'totales' del paso indicado. {} si no existe."""
    sub = "resumen" if paso == 1 else "totales"
    rp = db.query(ResultadoPaso).filter(
        ResultadoPaso.expediente_id == exp_id,
        ResultadoPaso.paso == paso,
        ResultadoPaso.subtipo == sub,
    ).first()
    if rp and isinstance(rp.datos, dict):
        return rp.datos
    return {}


def _cross_step_checks(db: Session, exp_id: int, paso: int) -> list[dict]:
    """
    Capa 2 — Checks de consistencia cruzada entre pasos.

    Compara totales del paso recién ejecutado contra pasos previos.
    Si hay desvíos significativos, devuelve alertas como warning (no bloquea).

    Formato de alerta: {nivel, titulo, mensaje}
      nivel: 'warning' | 'error' | 'info'
    """
    alertas: list[dict] = []

    def _ratio_check(a: float, b: float, tol_pct: float, titulo: str, ctx: str):
        """Si a/b se aparta más de tol_pct (ej. 0.05 = 5%) → warning."""
        if not a or not b:
            return
        ratio = a / b
        diff_pct = abs(ratio - 1) * 100
        if diff_pct > tol_pct * 100:
            alertas.append({
                "nivel": "warning",
                "titulo": titulo,
                "mensaje": (f"{ctx}: ratio {ratio:.3f} (diferencia {diff_pct:.1f}%, "
                            f"tolerancia {tol_pct*100:.0f}%). "
                            f"Valores: {a:,.0f} vs {b:,.0f}."),
            })

    # ── Paso 2 vs Paso 1 ──────────────────────────────────────────────
    if paso == 2:
        r1 = _get_resumen_paso(db, exp_id, 1)
        r2 = _get_resumen_paso(db, exp_id, 2)
        t1_gestion = r1.get("monto_total_gestion_usd")
        t2_facturado = r2.get("total_facturado_usd")
        _ratio_check(
            t2_facturado, t1_gestion, 0.05,
            "Paso 2 vs Paso 1: total facturado",
            "Paso 2.total_facturado vs Paso 1.total_gestion",
        )

    # ── Paso 3 vs Paso 2 ──────────────────────────────────────────────
    if paso == 3:
        r2 = _get_resumen_paso(db, exp_id, 2)
        r3 = _get_resumen_paso(db, exp_id, 3)
        t2_syngenta = r2.get("total_syngenta_usd")
        # Paso 3 puede tener varios totales — buscar el de gestion matched
        t3_gestion = r3.get("monto_gestion_total_usd") or r3.get("monto_g_total") \
                       or r3.get("total_gestion_usd")
        if t2_syngenta and t3_gestion:
            _ratio_check(
                t3_gestion, t2_syngenta, 0.10,
                "Paso 3 vs Paso 2: monto Syngenta",
                "Paso 3.gestion vs Paso 2.total_syngenta",
            )

    # ── Paso 4: sanity check de escala vs ARCA (compras vs ventas) ────
    if paso == 4:
        r1 = _get_resumen_paso(db, exp_id, 1)
        r4 = _get_resumen_paso(db, exp_id, 4)
        t1_arca = r1.get("monto_total_arca_usd")
        t4_compras = r4.get("total_compras_usd")
        if t1_arca and t4_compras:
            ratio = t4_compras / t1_arca
            # Las compras suelen ser 70-95% de las ventas del distribuidor.
            # Fuera de [0.3, 1.5] es sospechoso (compras 30% < ventas o 50% >).
            if ratio < 0.3 or ratio > 1.5:
                alertas.append({
                    "nivel": "warning",
                    "titulo": "Paso 4: escala de compras sospechosa vs ventas",
                    "mensaje": (
                        f"Compras (USD {t4_compras:,.0f}) representan {ratio*100:.0f}% "
                        f"de las ventas (USD {t1_arca:,.0f}). Fuera del rango típico "
                        f"30-150%. Verificar conversión de moneda y filtros."
                    ),
                })

    return alertas


# ── PASO 1 ────────────────────────────────────────────────────────────────────

@router.post("/1/ejecutar")
def ejecutar_paso1(
    exp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    exp = _get_exp(exp_id, db, current_user)

    # Múltiples archivos de Bajada de Gestión
    archivos_bajada = db.query(Archivo).filter(
        Archivo.expediente_id == exp_id,
        Archivo.tipo == TipoArchivo.BAJADA_GESTION,
    ).order_by(Archivo.created_at).all()
    if not archivos_bajada:
        raise HTTPException(400, "No hay archivos de Bajada de Gestión cargados")
    paths_bajada = []
    for a in archivos_bajada:
        if not os.path.exists(a.path):
            raise HTTPException(
                400,
                f"El archivo '{a.nombre_original}' se perdió del servidor "
                f"(probablemente por un redeploy). Volvé a subirlo desde el expediente."
            )
        paths_bajada.append((a.path, a.nombre_original))

    path_emitidos = _get_archivo_path(exp_id, TipoArchivo.COMPROBANTES_EMITIDOS, db)

    # Tipos de cambio: maestro global (Administración). Si no hay, error 400.
    from ...models.expediente import TipoCambioMaestro
    tc_count = db.query(TipoCambioMaestro).filter(TipoCambioMaestro.is_active == True).count()
    if tc_count == 0:
        raise HTTPException(
            400,
            "No hay tipos de cambio cargados. Pedile al administrador que los "
            "suba en Administración → Tipos de Cambio."
        )

    _t0 = time.time()
    _log_paso_inicio(db, exp_id, current_user.id, 1)
    import traceback
    try:
        resultado = paso1_service.ejecutar_paso1(paths_bajada, path_emitidos, exp_id, db=db)
    except Exception as e:
        _log_paso_error(db, exp_id, current_user.id, 1, _t0, e, traceback.format_exc())
        raise HTTPException(500, f"Error en Paso 1: {str(e)}")

    _save_resultado(db, exp_id, 1, "conciliacion", datos=resultado["conciliacion"])
    _save_resultado(db, exp_id, 1, "resumen", datos=resultado["resumen"])
    _save_resultado(db, exp_id, 1, "bajada_normalizada",
                    archivo_path=resultado["bajada_normalizada_path"])
    _save_resultado(db, exp_id, 1, "parser_diagnostico",
                    datos={"items": resultado.get("parser_diagnostico", [])})
    # Marcar completado ANTES de guardar la validación: si la validación falla
    # (API caída, timeout, etc.) el paso queda igualmente cerrado y se puede
    # avanzar al siguiente.
    _marcar_paso_completado(exp, 1, db)
    _safe_validar(db, exp_id, 1, lambda: resultado.get("validacion"))
    _log_paso_fin(db, exp_id, current_user.id, 1, _t0, contexto={
        "resumen": resultado.get("resumen"),
        "archivos_gestion": len(paths_bajada),
    })

    return {
        "resumen": resultado["resumen"],
        "conciliacion": resultado["conciliacion"][:500],  # Limitar para respuesta API
        "total_registros": len(resultado["conciliacion"]),
        "validacion": resultado.get("validacion"),
        "parser_diagnostico": resultado.get("parser_diagnostico", []),
    }


def _recalcular_resumen(conciliacion: list, resumen_original: dict) -> dict:
    """Recalcula el resumen desde las filas de conciliación actuales (post-anotaciones)."""
    total_arca      = sum(1 for r in conciliacion if (r.get("monto_usd_arca")     or 0) != 0)
    total_gestion   = sum(1 for r in conciliacion if (r.get("monto_usd_gestion")  or 0) != 0)
    solo_arca       = sum(1 for r in conciliacion if r.get("estado") == "SOLO_ARCA")
    solo_gestion    = sum(1 for r in conciliacion if r.get("estado") == "SOLO_GESTION")
    # OK: diferencia ≤ 1 USD y no es SOLO_*
    ok              = sum(1 for r in conciliacion
                         if r.get("estado") not in ("SOLO_ARCA", "SOLO_GESTION")
                         and abs(r.get("diferencia_usd") or 0) <= 1)
    con_diferencia  = sum(1 for r in conciliacion
                         if r.get("estado") not in ("SOLO_ARCA", "SOLO_GESTION")
                         and abs(r.get("diferencia_usd") or 0) > 1)
    monto_arca      = round(sum(r.get("monto_usd_arca")    or 0 for r in conciliacion), 2)
    monto_gestion   = round(sum(r.get("monto_usd_gestion") or 0 for r in conciliacion), 2)

    return {
        **resumen_original,          # conserva archivos_gestion, nombre_distribuidor, etc.
        "total_arca":               total_arca,
        "total_gestion":            total_gestion,
        "ok":                       ok,
        "con_diferencia":           con_diferencia,
        "solo_arca":                solo_arca,
        "solo_gestion":             solo_gestion,
        "monto_total_arca_usd":     monto_arca,
        "monto_total_gestion_usd":  monto_gestion,
    }


def _recalcular_alertas(conciliacion: list) -> list:
    """Regenera las alertas estructurales desde el estado actual de la conciliación."""
    alertas = []
    solo_g = [r for r in conciliacion if r.get("estado") == "SOLO_GESTION"]
    if solo_g:
        monto = sum(r.get("monto_usd_gestion") or 0 for r in solo_g)
        ejemplos = ", ".join(r.get("numero", "") for r in solo_g[:3])
        alertas.append({
            "nivel": "error",
            "titulo": f"{len(solo_g)} comprobante(s) en Gestión sin registro en ARCA",
            "detalle": (
                f"Monto total: US$ {monto:,.2f}. "
                f"Ejemplos: {ejemplos}. "
                "Esto no debería ocurrir: verificá que el archivo ARCA cubra el mismo período."
            ),
        })
    return alertas


@router.get("/1/resultado")
def resultado_paso1(
    exp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_exp(exp_id, db, current_user)
    res = db.query(ResultadoPaso).filter(
        ResultadoPaso.expediente_id == exp_id,
        ResultadoPaso.paso == 1,
    ).all()
    if not res:
        raise HTTPException(404, "Paso 1 no ejecutado aún")

    datos = {r.subtipo: r.datos for r in res}

    # Mergear anotaciones manuales en la conciliación
    anotaciones = db.query(AnotacionConciliacion).filter(
        AnotacionConciliacion.expediente_id == exp_id,
        AnotacionConciliacion.paso == 1,
    ).all()

    conciliacion = datos.get("conciliacion") or []

    if anotaciones and conciliacion:
        anot_idx = {a.comprobante_key: a for a in anotaciones}
        for row in conciliacion:
            key = row.get("key")
            if key and key in anot_idx:
                a = anot_idx[key]
                # Aplicar campos editados
                if a.cliente is not None:
                    row["cliente"] = a.cliente
                if a.monto_gestion_usd is not None:
                    row["monto_usd_gestion"] = a.monto_gestion_usd
                    row["diferencia_usd"] = round(
                        a.monto_gestion_usd - (row.get("monto_usd_arca") or 0), 2
                    )
                # Estado: MANUAL
                row["estado"] = "MANUAL"
                # Datos de anotación para el frontend
                row["anotacion"] = {
                    "es_agroquimico": a.es_agroquimico,
                    "producto": a.producto,
                    "cantidad": a.cantidad,
                    "unidad": a.unidad,
                    "cliente": a.cliente,
                    "monto_gestion_usd": a.monto_gestion_usd,
                }
        datos["conciliacion"] = conciliacion

    # Siempre recalcular resumen y alertas desde las filas actuales
    datos["resumen"] = _recalcular_resumen(conciliacion, datos.get("resumen") or {})
    datos["alertas"] = _recalcular_alertas(conciliacion)

    return datos


# ── Schema y endpoint de anotaciones ──────────────────────────────────────────

class AnotacionPayload(BaseModel):
    tipo: Optional[str] = None
    numero: Optional[str] = None
    fecha: Optional[str] = None
    cliente: Optional[str] = None
    monto_gestion_usd: Optional[float] = None
    es_agroquimico: Optional[bool] = None
    producto: Optional[str] = None
    cantidad: Optional[float] = None
    unidad: Optional[str] = None


@router.put("/1/anotaciones/{key:path}")
def guardar_anotacion(
    exp_id: int,
    key: str,
    payload: AnotacionPayload,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Crea o actualiza la anotación manual de un comprobante de la conciliación."""
    _get_exp(exp_id, db, current_user)

    anot = db.query(AnotacionConciliacion).filter(
        AnotacionConciliacion.expediente_id == exp_id,
        AnotacionConciliacion.paso == 1,
        AnotacionConciliacion.comprobante_key == key,
    ).first()

    if anot is None:
        anot = AnotacionConciliacion(
            expediente_id=exp_id,
            paso=1,
            comprobante_key=key,
        )
        db.add(anot)

    # Actualizar campos provistos
    for field in ("tipo", "numero", "fecha", "cliente", "monto_gestion_usd",
                  "es_agroquimico", "producto", "cantidad", "unidad"):
        val = getattr(payload, field)
        if val is not None:
            setattr(anot, field, val)

    db.commit()
    db.refresh(anot)
    return {"ok": True, "key": key}


@router.get("/1/anotaciones")
def listar_anotaciones(
    exp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_exp(exp_id, db, current_user)
    anotaciones = db.query(AnotacionConciliacion).filter(
        AnotacionConciliacion.expediente_id == exp_id,
        AnotacionConciliacion.paso == 1,
    ).all()
    return [
        {
            "key": a.comprobante_key,
            "tipo": a.tipo,
            "numero": a.numero,
            "fecha": a.fecha,
            "cliente": a.cliente,
            "monto_gestion_usd": a.monto_gestion_usd,
            "es_agroquimico": a.es_agroquimico,
            "producto": a.producto,
            "cantidad": a.cantidad,
            "unidad": a.unidad,
        }
        for a in anotaciones
    ]


# ── PASO 2 ────────────────────────────────────────────────────────────────────

@router.post("/2/ejecutar")
def ejecutar_paso2(
    exp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    exp = _get_exp(exp_id, db, current_user)

    # Verificar Paso 1 completado
    if 1 not in (exp.pasos_completados or []):
        raise HTTPException(400, "Completar Paso 1 antes de ejecutar Paso 2")

    # Obtener bajada normalizada
    bajada_res = db.query(ResultadoPaso).filter(
        ResultadoPaso.expediente_id == exp_id,
        ResultadoPaso.paso == 1,
        ResultadoPaso.subtipo == "bajada_normalizada",
    ).first()
    if not bajada_res or not bajada_res.archivo_path:
        raise HTTPException(400, "Bajada normalizada no disponible")

    # Maestro Syngenta
    maestro = db.query(MaestroSyngenta).filter(MaestroSyngenta.is_active == True).all()
    maestro_nombres = [m.nombre_estandar for m in maestro]

    # Clientes especiales (opcional)
    clientes_especiales = []
    arch_esp = db.query(Archivo).filter(
        Archivo.expediente_id == exp_id,
        Archivo.tipo == TipoArchivo.CLIENTES_ESPECIALES,
    ).first()
    if arch_esp:
        import pandas as pd
        df_esp = pd.read_excel(arch_esp.path, dtype=str)
        col = df_esp.columns[0]
        clientes_especiales = df_esp[col].dropna().str.strip().str.upper().tolist()

    _t0 = time.time()
    _log_paso_inicio(db, exp_id, current_user.id, 2)
    import traceback
    try:
        resultado = paso2_service.ejecutar_paso2(
            bajada_res.archivo_path,
            maestro_nombres,
            clientes_especiales,
            exp_id,
            exp.anio_analisis,
            db=db,
        )
    except Exception as e:
        _log_paso_error(db, exp_id, current_user.id, 2, _t0, e, traceback.format_exc())
        raise HTTPException(500, f"Error en Paso 2: {str(e)}")

    _save_resultado(db, exp_id, 2, "ranking_clientes", datos=resultado["ranking_clientes"])
    _save_resultado(db, exp_id, 2, "ranking_productos", datos=resultado["ranking_productos"])
    _save_resultado(db, exp_id, 2, "muestreo", datos=resultado["muestreo"])
    _save_resultado(db, exp_id, 2, "clasificacion", datos=resultado["clasificacion"])
    _save_resultado(db, exp_id, 2, "tabla_apertura", datos=resultado["tabla_apertura"])
    _save_resultado(db, exp_id, 2, "agroquimicos", archivo_path=resultado["agroquimicos_path"])
    _save_resultado(db, exp_id, 2, "syngenta", archivo_path=resultado["syngenta_path"])
    _save_resultado(db, exp_id, 2, "totales", datos=resultado["totales"])

    # Marcar completado primero — la validación que sigue es best-effort
    _marcar_paso_completado(exp, 2, db)
    validacion = _safe_validar(db, exp_id, 2, lambda: combinar_validacion(
        resultado.get("guardas", []),
        validar_paso(2, resultado["totales"], muestra=resultado["ranking_clientes"][:5]),
    ))
    _log_paso_fin(db, exp_id, current_user.id, 2, _t0, contexto={
        "totales": resultado.get("totales"),
    })

    return {
        "totales": resultado["totales"],
        "ranking_clientes_top10": resultado["ranking_clientes"][:10],
        "ranking_productos_top10": resultado["ranking_productos"][:10],
        "muestreo": resultado["muestreo"],
        "tabla_apertura": resultado["tabla_apertura"][:50],
        "validacion": validacion,
    }


@router.get("/2/resultado")
def resultado_paso2(
    exp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_exp(exp_id, db, current_user)
    res = db.query(ResultadoPaso).filter(
        ResultadoPaso.expediente_id == exp_id,
        ResultadoPaso.paso == 2,
    ).all()
    if not res:
        raise HTTPException(404, "Paso 2 no ejecutado aún")
    return {r.subtipo: r.datos for r in res}


# ── PASO 3 ────────────────────────────────────────────────────────────────────

@router.post("/3/ejecutar")
def ejecutar_paso3(
    exp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    exp = _get_exp(exp_id, db, current_user)

    if 2 not in (exp.pasos_completados or []):
        raise HTTPException(400, "Completar Paso 2 antes de ejecutar Paso 3")

    # Gestión a nivel línea para el cruce con CRM. Preferimos la bajada
    # NORMALIZADA del Paso 1 (gestión completa, antes de cualquier filtro de
    # IA) porque el Paso 3 ahora usa las marcas del CRM como ground truth
    # para identificar productos Syngenta — más confiable que la clasificación
    # con IA. Si no hay bajada_normalizada (expediente viejo), caemos al
    # archivo de Syngenta o de agroquímicos del Paso 2.
    def _archivo_paso(paso, subtipo):
        r = db.query(ResultadoPaso).filter(
            ResultadoPaso.expediente_id == exp_id,
            ResultadoPaso.paso == paso,
            ResultadoPaso.subtipo == subtipo,
        ).first()
        return r.archivo_path if r and r.archivo_path else None

    path_paso2 = (
        _archivo_paso(1, "bajada_normalizada")
        or _archivo_paso(2, "syngenta")
        or _archivo_paso(2, "agroquimicos")
    )
    if not path_paso2:
        raise HTTPException(
            400,
            "No se encontró la bajada de gestión normalizada. Re-ejecutá Paso 1.",
        )

    # CRM
    path_crm = _get_archivo_path(exp_id, TipoArchivo.CRM, db)

    # ── Anotaciones manuales del Paso 1 (solo las COMPLETAS) ────────────────
    # El auditor puede completar a mano comprobantes ARCA que la gestión no
    # detectó (status MANUAL en Paso 1). Para que cuenten en el cruce con CRM
    # tienen que estar bien clasificadas: agroquímico SÍ + producto + monto.
    # Las incompletas (cliente sin monto, sin clasificación) se ignoran para
    # no contaminar el cruce.
    anots_raw = db.query(AnotacionConciliacion).filter(
        AnotacionConciliacion.expediente_id == exp_id,
        AnotacionConciliacion.paso == 1,
        AnotacionConciliacion.es_agroquimico == True,
    ).all()
    anotaciones_manuales = []
    for a in anots_raw:
        if not a.producto or a.monto_gestion_usd is None or a.monto_gestion_usd == 0:
            continue
        # comprobante_key format: "FC-00001-00000034"
        parts = (a.comprobante_key or "").split("-", 2)
        tipo = parts[0] if len(parts) > 0 else (a.tipo or "FC")
        numero = "-".join(parts[1:]) if len(parts) > 1 else (a.numero or "")
        anotaciones_manuales.append({
            "tipo_comprobante": tipo,
            "tipo": tipo,
            "numero_comprobante": numero,
            "numero": numero,
            "fecha": a.fecha or "",
            "cliente": a.cliente or "",
            "cuit_cliente": "",
            "articulo": a.producto,
            "cantidad": a.cantidad or 0,
            "monto_usd": a.monto_gestion_usd,
        })

    _t0 = time.time()
    _log_paso_inicio(db, exp_id, current_user.id, 3)
    # Capturamos TODO con traceback completo para que el toast del frontend muestre
    # el detalle real en vez de "Internal Server Error" genérico.
    import traceback
    try:
        resultado = paso3_service.ejecutar_paso3(
            path_paso2, path_crm, exp_id,
            cuit_distribuidor=exp.cuit_distribuidor,
            nombre_distribuidor=exp.nombre_distribuidor,
            anotaciones_manuales=anotaciones_manuales,
        )
        _save_resultado(db, exp_id, 3, "conciliacion", datos=resultado["conciliacion"])
        _save_resultado(db, exp_id, 3, "resumen", datos=resultado["resumen"])
        _save_resultado(db, exp_id, 3, "parser_diagnostico",
                        datos={"items": resultado.get("parser_diagnostico", [])})
    except HTTPException:
        raise
    except Exception as e:
        tb = traceback.format_exc()
        print(f"[PASO 3] EXCEPTION: {tb}")
        _log_paso_error(db, exp_id, current_user.id, 3, _t0, e, tb)
        raise HTTPException(500, f"Error en Paso 3: {type(e).__name__}: {str(e) or '(sin mensaje)'}")

    # Marcar completado primero — la validación que sigue es best-effort
    _marcar_paso_completado(exp, 3, db)
    validacion = _safe_validar(db, exp_id, 3, lambda: validar_paso(
        3, resultado["resumen"], muestra=resultado["conciliacion"][:5],
    ))
    _log_paso_fin(db, exp_id, current_user.id, 3, _t0, contexto={
        "resumen": resultado.get("resumen"),
        "filtro_distribuidor": resultado.get("filtro_distribuidor"),
        "anotaciones_manuales": len(anotaciones_manuales),
    })

    return {**resultado, "validacion": validacion}


@router.get("/3/resultado")
def resultado_paso3(
    exp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_exp(exp_id, db, current_user)
    res = db.query(ResultadoPaso).filter(
        ResultadoPaso.expediente_id == exp_id,
        ResultadoPaso.paso == 3,
    ).all()
    if not res:
        raise HTTPException(404, "Paso 3 no ejecutado aún")
    return {r.subtipo: r.datos for r in res}


# ── PASO 4 ────────────────────────────────────────────────────────────────────

@router.post("/4/ejecutar")
def ejecutar_paso4(
    exp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    exp = _get_exp(exp_id, db, current_user)

    path_recibidos = _get_archivo_path(exp_id, TipoArchivo.COMPROBANTES_RECIBIDOS, db)

    # Tipos de cambio: maestro global (Administración). Si no hay, error 400.
    from ...models.expediente import TipoCambioMaestro
    tc_count = db.query(TipoCambioMaestro).filter(TipoCambioMaestro.is_active == True).count()
    if tc_count == 0:
        raise HTTPException(
            400,
            "No hay tipos de cambio cargados. Pedile al administrador que los "
            "suba en Administración → Tipos de Cambio."
        )

    # Archivo de proveedores: configura cuáles abrir en FC/NC/ND ("ABRIR")
    # y cuáles incluir siempre en el top 90% aunque sean chicos ("INCLUIR").
    # Formato flexible — ver paso4_service.leer_archivo_proveedores.
    proveedores_config = {"por_cuit": {}, "por_nombre": {}}
    arch_prov = db.query(Archivo).filter(
        Archivo.expediente_id == exp_id,
        Archivo.tipo == TipoArchivo.PROVEEDORES_APERTURA,
    ).first()
    if arch_prov:
        proveedores_config = paso4_service.leer_archivo_proveedores(arch_prov.path)

    _t0 = time.time()
    _log_paso_inicio(db, exp_id, current_user.id, 4)
    import traceback
    try:
        resultado = paso4_service.ejecutar_paso4(
            path_recibidos, proveedores_config, exp.anio_analisis, db=db
        )
    except Exception as e:
        _log_paso_error(db, exp_id, current_user.id, 4, _t0, e, traceback.format_exc())
        raise HTTPException(500, f"Error en Paso 4: {str(e)}")

    _save_resultado(db, exp_id, 4, "resumen", datos=resultado["resumen"])
    _save_resultado(db, exp_id, 4, "totales", datos=resultado["totales"])
    _save_resultado(db, exp_id, 4, "parser_diagnostico",
                    datos={"items": resultado.get("parser_diagnostico", [])})

    # Marcar completado primero — la validación que sigue es best-effort
    _marcar_paso_completado(exp, 4, db)
    validacion = _safe_validar(db, exp_id, 4, lambda: validar_paso(
        4, resultado["totales"], muestra=resultado["resumen"][:5],
    ))
    _log_paso_fin(db, exp_id, current_user.id, 4, _t0, contexto={
        "totales": resultado.get("totales"),
    })

    return {
        "totales": resultado["totales"],
        "resumen_top20": resultado["resumen"][:20],
        "parser_diagnostico": resultado.get("parser_diagnostico", []),
        "validacion": validacion,
    }


@router.get("/4/resultado")
def resultado_paso4(
    exp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_exp(exp_id, db, current_user)
    res = db.query(ResultadoPaso).filter(
        ResultadoPaso.expediente_id == exp_id,
        ResultadoPaso.paso == 4,
    ).all()
    if not res:
        raise HTTPException(404, "Paso 4 no ejecutado aún")
    return {r.subtipo: r.datos for r in res}


# ── PASO 5 ────────────────────────────────────────────────────────────────────

@router.post("/5/ejecutar")
def ejecutar_paso5(
    exp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    exp = _get_exp(exp_id, db, current_user)

    reqs = [2, 3, 4]
    faltantes = [p for p in reqs if p not in (exp.pasos_completados or [])]
    if faltantes:
        raise HTTPException(400, f"Completar pasos {faltantes} antes del Paso 5")

    # Cargar datos de pasos anteriores
    def get_datos(paso, subtipo):
        r = db.query(ResultadoPaso).filter(
            ResultadoPaso.expediente_id == exp_id,
            ResultadoPaso.paso == paso,
            ResultadoPaso.subtipo == subtipo,
        ).first()
        return r.datos if r else []

    resumen_compras = get_datos(4, "resumen")
    tabla_apertura = get_datos(2, "tabla_apertura")
    conciliacion_crm = get_datos(3, "conciliacion")

    expediente_info = {
        "nombre_distribuidor": exp.nombre_distribuidor,
        "cuit_distribuidor": exp.cuit_distribuidor,
        "anio_analisis": exp.anio_analisis,
    }

    _t0 = time.time()
    _log_paso_inicio(db, exp_id, current_user.id, 5)
    import traceback
    try:
        resultado = paso5_service.ejecutar_paso5(
            resumen_compras, tabla_apertura, conciliacion_crm,
            expediente_info, exp_id
        )
    except Exception as e:
        _log_paso_error(db, exp_id, current_user.id, 5, _t0, e, traceback.format_exc())
        raise HTTPException(500, f"Error en Paso 5: {str(e)}")

    _save_resultado(db, exp_id, 5, "informe", archivo_path=resultado["excel_path"])
    _save_resultado(db, exp_id, 5, "totales", datos=resultado["totales"])

    # Marcar completado primero — la validación que sigue es best-effort
    _marcar_paso_completado(exp, 5, db)
    validacion = _safe_validar(db, exp_id, 5, lambda: validar_paso(5, resultado["totales"]))
    _log_paso_fin(db, exp_id, current_user.id, 5, _t0, contexto={"totales": resultado.get("totales")})

    return {**resultado, "validacion": validacion}


@router.get("/5/descargar")
def descargar_paso5(
    exp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_exp(exp_id, db, current_user)
    res = db.query(ResultadoPaso).filter(
        ResultadoPaso.expediente_id == exp_id,
        ResultadoPaso.paso == 5,
        ResultadoPaso.subtipo == "informe",
    ).first()
    if not res or not res.archivo_path:
        raise HTTPException(404, "Informe no generado aún")
    return FileResponse(
        res.archivo_path,
        filename=f"OAG_Informe_Paso5_Exp{exp_id}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── PASO 6 ────────────────────────────────────────────────────────────────────

@router.post("/6/ejecutar")
def ejecutar_paso6(
    exp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    exp = _get_exp(exp_id, db, current_user)

    if 5 not in (exp.pasos_completados or []):
        raise HTTPException(400, "Completar Paso 5 antes del Paso 6")

    def get_datos(paso, subtipo):
        r = db.query(ResultadoPaso).filter(
            ResultadoPaso.expediente_id == exp_id,
            ResultadoPaso.paso == paso,
            ResultadoPaso.subtipo == subtipo,
        ).first()
        return r.datos if r else []

    resumen_compras = get_datos(4, "resumen")
    tabla_apertura = get_datos(2, "tabla_apertura")
    conciliacion_crm = get_datos(3, "conciliacion")

    glosario = db.query(Glosario).filter(Glosario.is_active == True).all()
    glosario_list = [{"nombre_original": g.nombre_original, "nombre_estandar": g.nombre_estandar} for g in glosario]

    expediente_info = {
        "nombre_distribuidor": exp.nombre_distribuidor,
        "cuit_distribuidor": exp.cuit_distribuidor,
        "anio_analisis": exp.anio_analisis,
    }

    _t0 = time.time()
    _log_paso_inicio(db, exp_id, current_user.id, 6)
    import traceback
    try:
        resultado = paso6_service.ejecutar_paso6(
            resumen_compras, tabla_apertura, conciliacion_crm,
            glosario_list, expediente_info, exp_id
        )
    except Exception as e:
        _log_paso_error(db, exp_id, current_user.id, 6, _t0, e, traceback.format_exc())
        raise HTTPException(500, f"Error en Paso 6: {str(e)}")

    _save_resultado(db, exp_id, 6, "informe", archivo_path=resultado["excel_path"])
    _save_resultado(db, exp_id, 6, "totales", datos=resultado["totales"])

    # Marcar completado primero — la validación que sigue es best-effort
    _marcar_paso_completado(exp, 6, db)
    validacion = _safe_validar(db, exp_id, 6, lambda: validar_paso(6, resultado["totales"]))
    _log_paso_fin(db, exp_id, current_user.id, 6, _t0, contexto={"totales": resultado.get("totales")})

    return {**resultado, "validacion": validacion}


@router.get("/6/descargar")
def descargar_paso6(
    exp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_exp(exp_id, db, current_user)
    res = db.query(ResultadoPaso).filter(
        ResultadoPaso.expediente_id == exp_id,
        ResultadoPaso.paso == 6,
        ResultadoPaso.subtipo == "informe",
    ).first()
    if not res or not res.archivo_path:
        raise HTTPException(404, "Informe no generado aún")
    return FileResponse(
        res.archivo_path,
        filename=f"OAG_Informe_Final_Exp{exp_id}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── EXPORT DE OUTPUT POR PASO (debug / soporte) ───────────────────────────────

@router.get("/{paso}/exportar")
def exportar_paso(
    exp_id: int,
    paso: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Exporta TODO lo que genera un paso como un .zip:
      - _meta.json: metadatos del expediente y del paso
      - datos.json: todos los subtipos con datos tabulares (conciliación,
        resumen, totales, validación, etc.)
      - archivos: cualquier .xlsx/.csv que el paso haya producido (bajada
        normalizada, agroquímicos, syngenta, informes, etc.)
    Útil para reportar bugs sin tener que copiar pantallas — adjuntar el zip
    a un ticket y el dev tiene todo lo necesario para reproducir.
    """
    if paso < 1 or paso > 6:
        raise HTTPException(400, "Paso debe ser 1-6")
    exp = _get_exp(exp_id, db, current_user)
    resultados = (
        db.query(ResultadoPaso)
        .filter(ResultadoPaso.expediente_id == exp_id, ResultadoPaso.paso == paso)
        .all()
    )
    if not resultados:
        raise HTTPException(404, f"Paso {paso} no ejecutado aún")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        # Metadata del expediente y del paso
        meta = {
            "expediente_id": exp_id,
            "distribuidor": exp.nombre_distribuidor,
            "cuit": exp.cuit_distribuidor,
            "anio_analisis": exp.anio_analisis,
            "paso": paso,
            "exportado": datetime.utcnow().isoformat() + "Z",
            "exportado_por": current_user.email,
            "subtipos_presentes": [r.subtipo for r in resultados],
        }
        z.writestr("_meta.json",
                   json.dumps(meta, ensure_ascii=False, indent=2, default=str))

        # Datos tabulares
        datos_dict = {}
        for r in resultados:
            if r.datos is not None:
                datos_dict[r.subtipo] = r.datos
        if datos_dict:
            z.writestr(
                "datos.json",
                json.dumps(datos_dict, ensure_ascii=False, indent=2, default=str),
            )

        # Archivos físicos (xlsx, csv) generados por el paso
        for r in resultados:
            if r.archivo_path and os.path.exists(r.archivo_path):
                arcname = f"archivos/{r.subtipo}_{os.path.basename(r.archivo_path)}"
                try:
                    z.write(r.archivo_path, arcname)
                except Exception as e:
                    z.writestr(f"archivos/_ERROR_{r.subtipo}.txt",
                               f"No se pudo adjuntar {r.archivo_path}: {e}")

    buf.seek(0)
    fname = (
        f"OAG_paso{paso}_exp{exp_id}_"
        f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.zip"
    )
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── EXPORT COMPLETO (todos los pasos en un solo Excel) ────────────────────────

@router.get("/exportar-completo")
def exportar_completo(
    exp_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Genera un único Excel con una hoja por paso ejecutado:
      - Métricas: resumen numérico de todos los pasos
      - P1 Conciliación ARCA: tipo/número/fecha/cliente/estado/montos
      - P2 Clasificación: tabla de apertura mensual por cliente
      - P2 Productos: ranking de productos con clasificación agro/Syngenta
      - P3 Cruce CRM: conciliación gestión vs CRM por marca
      - P4 Compras: detalle por proveedor con totales mensuales
    Diseñado para comparar fácilmente con el Excel del auditor humano.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    exp = _get_exp(exp_id, db, current_user)

    # Leer todos los resultados de una vez
    todos = (
        db.query(ResultadoPaso)
        .filter(ResultadoPaso.expediente_id == exp_id)
        .all()
    )
    data: dict[tuple, any] = {(r.paso, r.subtipo): r.datos for r in todos}

    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja vacía default

    # ── Estilos reutilizables ──────────────────────────────────────────────────
    H_FILL  = PatternFill("solid", fgColor="1C2B3A")
    H_FONT  = Font(color="FFFFFF", bold=True, size=9)
    H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    D_FONT  = Font(size=9)
    D_ALIGN = Alignment(vertical="center")
    Z_FILL  = PatternFill("solid", fgColor="F0F4F8")
    OK_FILL = PatternFill("solid", fgColor="E6F4EA")
    ER_FILL = PatternFill("solid", fgColor="FDECEA")
    WA_FILL = PatternFill("solid", fgColor="FFF8E1")
    BORDER  = Border(
        left=Side(style="thin", color="DDE1E7"),
        right=Side(style="thin", color="DDE1E7"),
        top=Side(style="thin", color="DDE1E7"),
        bottom=Side(style="thin", color="DDE1E7"),
    )

    def _write_header(ws, cols: list[str]):
        for i, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=i, value=col)
            c.fill = H_FILL; c.font = H_FONT
            c.alignment = H_ALIGN; c.border = BORDER

    def _write_row(ws, row_num: int, vals: list, fill=None):
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=i, value=v)
            c.font = D_FONT; c.alignment = D_ALIGN; c.border = BORDER
            if fill:
                c.fill = fill
            elif row_num % 2 == 0:
                c.fill = Z_FILL

    def _autowidth(ws, max_w=50):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, max_w)

    def _freeze(ws):
        ws.freeze_panes = "A2"

    # ── Hoja: Métricas ────────────────────────────────────────────────────────
    ws_m = wb.create_sheet("Métricas")
    ws_m.sheet_properties.tabColor = "1A4A8A"

    # Header de info
    ws_m["A1"] = f"DS: {exp.nombre_distribuidor}"
    ws_m["A1"].font = Font(bold=True, size=11)
    ws_m["A2"] = f"CUIT: {exp.cuit_distribuidor or '—'}   Año: {exp.anio_analisis or '—'}"
    ws_m["A2"].font = Font(size=9, color="666666")
    ws_m["A3"] = f"Exportado: {datetime.utcnow().strftime('%d/%m/%Y %H:%M')} UTC"
    ws_m["A3"].font = Font(size=9, color="666666")
    ws_m.row_dimensions[1].height = 20
    ws_m.append([])  # fila 4 vacía

    _write_header(ws_m, ["Paso", "Métrica", "Valor"])
    ws_m.column_dimensions["A"].width = 10
    ws_m.column_dimensions["B"].width = 35
    ws_m.column_dimensions["C"].width = 22

    row = 6
    def _metric_row(paso_num, label, value):
        nonlocal row
        ws_m.cell(row=row, column=1, value=f"Paso {paso_num}").font = D_FONT
        ws_m.cell(row=row, column=2, value=label).font = D_FONT
        ws_m.cell(row=row, column=3, value=value).font = D_FONT
        for col in range(1, 4):
            ws_m.cell(row=row, column=col).border = BORDER
            if row % 2 == 0:
                ws_m.cell(row=row, column=col).fill = Z_FILL
        row += 1

    # Paso 1
    r1 = data.get((1, "resumen")) or {}
    if r1:
        _metric_row(1, "ARCA total",           r1.get("total_arca"))
        _metric_row(1, "Gestión total",         r1.get("total_gestion"))
        _metric_row(1, "Matcheados OK",         r1.get("ok"))
        _metric_row(1, "Solo ARCA",             r1.get("solo_arca"))
        _metric_row(1, "Solo Gestión",          r1.get("solo_gestion"))
        _metric_row(1, "Internos",              r1.get("internos"))
        _metric_row(1, "Monto ARCA (USD)",      round(r1.get("monto_total_arca_usd") or 0))
        _metric_row(1, "Monto Gestión (USD)",   round(r1.get("monto_total_gestion_usd") or 0))
    # Paso 2
    t2 = data.get((2, "totales")) or {}
    if t2:
        _metric_row(2, "Total facturado (USD)", round(t2.get("total_facturado_usd") or 0))
        _metric_row(2, "Total Syngenta (USD)",  round(t2.get("total_syngenta_usd") or 0))
        _metric_row(2, "Total agroquímicos (USD)", round(t2.get("total_agroquimicos_usd") or 0))
        fac = t2.get("total_facturado_usd") or 0
        syn = t2.get("total_syngenta_usd") or 0
        _metric_row(2, "% Syngenta",            f"{syn/fac*100:.1f}%" if fac else "—")
    # Paso 3
    r3 = data.get((3, "resumen")) or {}
    if r3:
        _metric_row(3, "Total filas CRM",       r3.get("total"))
        _metric_row(3, "OK cruzados",           r3.get("ok"))
        _metric_row(3, "Solo CRM",              r3.get("solo_crm"))
        _metric_row(3, "Solo Gestión",          r3.get("solo_gestion"))
        _metric_row(3, "Monto Gestión (USD)",   round(r3.get("monto_gestion_total_usd") or 0))
        _metric_row(3, "Monto CRM (USD)",       round(r3.get("monto_crm_total_usd") or 0))
    # Paso 4
    t4 = data.get((4, "totales")) or {}
    if t4:
        _metric_row(4, "Total compras (USD)",   round(t4.get("total_compras_usd") or 0))

    # ── Hoja: P1 Conciliación ─────────────────────────────────────────────────
    conc1 = data.get((1, "conciliacion")) or []
    if conc1:
        ws1 = wb.create_sheet("P1 Conciliación")
        COLS1 = ["Tipo", "Número", "Fecha", "Cliente", "Estado",
                 "Monto ARCA (USD)", "Monto Gestión (USD)", "Diferencia (USD)", "Productos"]
        _write_header(ws1, COLS1)
        _freeze(ws1)
        for i, row_d in enumerate(conc1, 2):
            estado = str(row_d.get("estado") or "")
            fill = (OK_FILL if estado == "OK"
                    else ER_FILL if estado in ("SOLO_ARCA", "SOLO_GESTION")
                    else WA_FILL if estado == "DIFERENCIA"
                    else None)
            _write_row(ws1, i, [
                row_d.get("tipo"),
                row_d.get("numero"),
                row_d.get("fecha"),
                row_d.get("cliente"),
                estado,
                round(row_d.get("monto_usd_arca") or 0, 2),
                round(row_d.get("monto_usd_gestion") or 0, 2),
                round(row_d.get("diferencia_usd") or 0, 2),
                (row_d.get("productos") or "")[:80],
            ], fill=fill if fill else None)
        _autowidth(ws1)
        ws1.sheet_properties.tabColor = "2E7D32"

    # ── Hoja: P2 Tabla Apertura ───────────────────────────────────────────────
    tabla2 = data.get((2, "tabla_apertura")) or []
    if tabla2:
        ws2a = wb.create_sheet("P2 Apertura mensual")
        if tabla2:
            cols2a = list(tabla2[0].keys())
            _write_header(ws2a, cols2a)
            _freeze(ws2a)
            for i, row_d in enumerate(tabla2, 2):
                _write_row(ws2a, i, [row_d.get(c) for c in cols2a])
            _autowidth(ws2a)
        ws2a.sheet_properties.tabColor = "1565C0"

    # ── Hoja: P2 Productos ────────────────────────────────────────────────────
    clas2 = data.get((2, "clasificacion")) or []
    rank2 = data.get((2, "ranking_productos")) or []
    prods_data = clas2 if clas2 else rank2
    if prods_data:
        ws2p = wb.create_sheet("P2 Productos")
        if clas2:
            COLS2P = ["Producto", "Agroquímico", "Syngenta", "Justificación"]
            _write_header(ws2p, COLS2P)
            _freeze(ws2p)
            for i, row_d in enumerate(clas2, 2):
                syng = str(row_d.get("syngenta") or "")
                fill = (OK_FILL  if syng == "SI"
                        else WA_FILL if syng == "REVISAR"
                        else None)
                _write_row(ws2p, i, [
                    row_d.get("producto"),
                    row_d.get("agroquimico"),
                    syng,
                    (row_d.get("justificacion") or "")[:100],
                ], fill=fill)
        else:
            cols2p = list(rank2[0].keys()) if rank2 else []
            _write_header(ws2p, cols2p)
            _freeze(ws2p)
            for i, row_d in enumerate(rank2, 2):
                _write_row(ws2p, i, [row_d.get(c) for c in cols2p])
        _autowidth(ws2p)
        ws2p.sheet_properties.tabColor = "1565C0"

    # ── Hoja: P3 Cruce CRM ────────────────────────────────────────────────────
    conc3 = data.get((3, "conciliacion")) or []
    if conc3:
        ws3 = wb.create_sheet("P3 Cruce CRM")
        # Detectar columnas disponibles dinámicamente
        sample = conc3[0] if conc3 else {}
        PRIORITY3 = ["marca_crm", "marca", "producto_crm", "cliente_crm", "cliente",
                     "fecha_crm", "numero_crm", "tipo_crm", "estado",
                     "monto_gestion", "monto_usd_gestion", "monto_crm", "monto_usd_crm",
                     "diferencia", "diferencia_usd", "justificacion"]
        cols3 = [c for c in PRIORITY3 if c in sample] or list(sample.keys())[:15]
        LABELS3 = {
            "marca_crm": "Marca", "marca": "Marca", "producto_crm": "Producto CRM",
            "cliente_crm": "Cliente CRM", "cliente": "Cliente",
            "fecha_crm": "Fecha", "numero_crm": "Número", "tipo_crm": "Tipo",
            "estado": "Estado",
            "monto_gestion": "Monto Gestión (USD)", "monto_usd_gestion": "Monto Gestión (USD)",
            "monto_crm": "Monto CRM (USD)", "monto_usd_crm": "Monto CRM (USD)",
            "diferencia": "Diferencia (USD)", "diferencia_usd": "Diferencia (USD)",
            "justificacion": "Justificación IA",
        }
        _write_header(ws3, [LABELS3.get(c, c) for c in cols3])
        _freeze(ws3)
        for i, row_d in enumerate(conc3, 2):
            estado = str(row_d.get("estado") or "")
            fill = (OK_FILL if estado == "OK"
                    else ER_FILL if estado in ("SOLO_CRM", "SOLO_GESTION")
                    else WA_FILL if "DIF" in estado.upper()
                    else None)
            vals3 = []
            for c in cols3:
                v = row_d.get(c)
                if isinstance(v, float): v = round(v, 2)
                if isinstance(v, str) and len(v) > 100: v = v[:100]
                vals3.append(v)
            _write_row(ws3, i, vals3, fill=fill)
        _autowidth(ws3)
        ws3.sheet_properties.tabColor = "6A1B9A"

    # ── Hoja: P4 Compras ──────────────────────────────────────────────────────
    res4 = data.get((4, "resumen")) or []
    if res4:
        ws4 = wb.create_sheet("P4 Compras")
        sample4 = res4[0] if res4 else {}
        # Columnas fijas primero, luego dinámicas
        FIXED4 = ["proveedor", "nombre_proveedor", "cuit", "cuit_proveedor",
                  "Total", "total_usd", "categoria"]
        avail4 = list(sample4.keys())
        # Primero las fijas que existan, luego el resto
        cols4 = [c for c in FIXED4 if c in sample4]
        cols4 += [c for c in avail4 if c not in cols4]
        LABELS4 = {
            "proveedor": "Proveedor", "nombre_proveedor": "Proveedor",
            "cuit": "CUIT", "cuit_proveedor": "CUIT",
            "Total": "Total (USD)", "total_usd": "Total (USD)",
            "categoria": "Categoría",
        }
        _write_header(ws4, [LABELS4.get(c, c) for c in cols4])
        _freeze(ws4)
        for i, row_d in enumerate(res4, 2):
            vals4 = []
            for c in cols4:
                v = row_d.get(c)
                if isinstance(v, float): v = round(v, 2)
                vals4.append(v)
            _write_row(ws4, i, vals4)
        _autowidth(ws4)
        ws4.sheet_properties.tabColor = "E65100"

    # ── Serializar ────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    dist_slug = (exp.nombre_distribuidor or f"exp{exp_id}").replace(" ", "_")[:30]
    fname = f"OAG_{dist_slug}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
