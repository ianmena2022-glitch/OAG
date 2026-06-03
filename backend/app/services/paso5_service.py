"""
Paso 5 — Informe Ejecutivo (Compras y Ventas)
Genera Excel + PDF con 3 anexos: Resumen Compras (top 90%), Venta Agroquímicos, Cruce CRM.
"""
import os
import io
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from ..core.config import settings

# Paleta OAG
COLOR_HEADER = "1C2B3A"   # Azul carbón oscuro
COLOR_HEADER_FONT = "FFFFFF"
COLOR_SYNGENTA = "1A4A8A"  # Azul corporativo
COLOR_SYNGENTA_FONT = "FFFFFF"
COLOR_ZEBRA = "F0F4F8"
COLOR_DIFERENCIA = "FDECEA"
COLOR_OK = "EAF4EA"
COLOR_BORDER = "DDE1E7"
COLOR_TOTAL = "E8EDF3"

MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


def ejecutar_paso5(
    resumen_compras: List[Dict],
    tabla_apertura: List[Dict],
    conciliacion_crm: List[Dict],
    expediente_info: Dict,
    expediente_id: int,
) -> dict:
    """
    Genera informe ejecutivo sin glosario (Paso 5).
    """
    upload_dir = os.path.join(settings.UPLOAD_DIR, str(expediente_id))
    os.makedirs(upload_dir, exist_ok=True)

    # Filtrar top 90% compras
    resumen_top90 = _filtrar_top90(resumen_compras)

    # Generar Excel
    path_excel = os.path.join(upload_dir, "informe_paso5.xlsx")
    _generar_excel(resumen_top90, tabla_apertura, conciliacion_crm, expediente_info, path_excel)

    return {
        "excel_path": path_excel,
        "totales": {
            "proveedores_incluidos": len(resumen_top90),
            "total_proveedores": len(resumen_compras),
            "lineas_apertura": len(tabla_apertura),
            "lineas_crm": len(conciliacion_crm),
        },
    }


def _filtrar_top90(resumen: List[Dict]) -> List[Dict]:
    """
    Retorna los proveedores a incluir en el informe:
      - Siempre los marcados como ABRIR o INCLUIR en el archivo de proveedores
        (aunque su monto sea chico)
      - Más los demás proveedores hasta cubrir el 90% del total
    """
    if not resumen:
        return []

    siempre = [r for r in resumen if r.get("categoria") in ("ABRIR", "INCLUIR")]
    resto = [r for r in resumen if r.get("categoria") not in ("ABRIR", "INCLUIR")]

    total_resto = sum(abs(r.get("Total", 0)) for r in resto)
    if total_resto == 0:
        # No hay otros — devolver solo los siempre (o todo si tampoco hay siempre)
        return siempre if siempre else resumen

    ordenado = sorted(resto, key=lambda x: abs(x.get("Total", 0)), reverse=True)
    acumulado = 0
    top = []
    for r in ordenado:
        top.append(r)
        acumulado += abs(r.get("Total", 0))
        if acumulado / total_resto >= 0.90:
            break

    # Ordenar el resultado final por Total descendente, conservando todos
    return sorted(siempre + top, key=lambda x: abs(x.get("Total", 0)), reverse=True)


def _estilo_header(ws, row: int, cols: list, color: str = COLOR_HEADER, font_color: str = COLOR_HEADER_FONT):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font = Font(name="Calibri", bold=True, color=font_color, size=10)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=COLOR_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, value in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border


def _estilo_celda(ws, row: int, col: int, value, is_zebra: bool = False, formato=None, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", size=9, bold=bold)
    cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left",
                                vertical="center")
    if is_zebra:
        cell.fill = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")
    if formato:
        cell.number_format = formato
    thin = Side(style="thin", color="E0E4E8")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell


def _titulo_hoja(ws, titulo: str, subtitulo: str = ""):
    ws.merge_cells("A1:Z1")
    cell = ws["A1"]
    cell.value = titulo
    cell.font = Font(name="Calibri", bold=True, size=13, color=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    if subtitulo:
        ws.merge_cells("A2:Z2")
        sub = ws["A2"]
        sub.value = subtitulo
        sub.font = Font(name="Calibri", size=9, color="888888")
        sub.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 14


def _generar_excel(
    resumen_compras: List[Dict],
    tabla_apertura: List[Dict],
    conciliacion_crm: List[Dict],
    expediente_info: Dict,
    path: str,
):
    wb = Workbook()
    wb.remove(wb.active)

    # Portada
    _hoja_portada(wb, expediente_info)
    # Anexo 1: Resumen Compras
    _hoja_resumen_compras(wb, resumen_compras)
    # Anexo 2: Venta Agroquímicos
    _hoja_tabla_apertura(wb, tabla_apertura)
    # Anexo 3: Cruce CRM
    _hoja_conciliacion_crm(wb, conciliacion_crm)

    wb.save(path)


def _hoja_portada(wb: Workbook, info: Dict):
    ws = wb.create_sheet("Portada")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40

    # Header azul
    for row in range(1, 8):
        for col in range(1, 20):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

    ws.merge_cells("B2:I4")
    title = ws["B2"]
    title.value = "OGSA — SISTEMA DE AUDITORÍAS COMERCIALES"
    title.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    title.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B5:I6")
    sub = ws["B5"]
    sub.value = "Informe de Compras y Ventas"
    sub.font = Font(name="Calibri", size=11, color="AABDD4")
    sub.alignment = Alignment(horizontal="left", vertical="center")

    # Datos del expediente
    datos = [
        ("Distribuidor:", info.get("nombre_distribuidor", "")),
        ("CUIT:", info.get("cuit_distribuidor", "")),
        ("Período bajo análisis:", str(info.get("anio_analisis", ""))),
        ("Índice:", "Anexo I — Resumen de Compras | Anexo II — Venta de Agroquímicos | Anexo III — Cruce CRM"),
    ]
    for i, (label, value) in enumerate(datos, start=10):
        ws.cell(row=i, column=2, value=label).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=i, column=3, value=value).font = Font(name="Calibri", size=10)
        ws.row_dimensions[i].height = 16


def _hoja_resumen_compras(wb: Workbook, resumen: List[Dict]):
    ws = wb.create_sheet("Anexo I — Compras")
    ws.sheet_view.showGridLines = False

    _titulo_hoja(ws, "ANEXO I — RESUMEN DE COMPRAS", "Compras por proveedor — Top 90% del total (en USD)")

    # Orden: Proveedor | CUIT | TOTAL USD | Enero..Diciembre
    headers = ["Proveedor", "CUIT", "TOTAL USD"] + MESES
    _estilo_header(ws, 4, headers)

    for i, row in enumerate(resumen):
        r = i + 5
        zebra = i % 2 == 0
        _estilo_celda(ws, r, 1, row.get("nombre_proveedor", ""), zebra)
        _estilo_celda(ws, r, 2, row.get("cuit_proveedor", ""), zebra)
        total = row.get("Total", 0)
        c = _estilo_celda(ws, r, 3, total, zebra, '#,##0.00', bold=True)
        if total < 0:
            c.font = Font(name="Calibri", size=9, bold=True, color="C0392B")
        for m_idx, mes in enumerate(MESES):
            val = row.get(mes, 0)
            _estilo_celda(ws, r, 4 + m_idx, val if val else None, zebra, '#,##0.00')

    # Totalizadora
    total_row = len(resumen) + 5
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Calibri", size=10)
    gran_total = sum(r.get("Total", 0) for r in resumen)
    ws.cell(row=total_row, column=3, value=round(gran_total, 2)).number_format = '#,##0.00'
    ws.cell(row=total_row, column=3).font = Font(bold=True, name="Calibri", size=10)
    for m_idx in range(12):
        col_vals = [r.get(MESES[m_idx], 0) for r in resumen]
        ws.cell(row=total_row, column=4 + m_idx, value=round(sum(col_vals), 2)).number_format = '#,##0.00'

    # Anchos de columna
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14   # TOTAL USD
    for col in range(4, 16):
        ws.column_dimensions[get_column_letter(col)].width = 13


def _hoja_tabla_apertura(wb: Workbook, tabla: List[Dict]):
    ws = wb.create_sheet("Anexo II — Agroquímicos")
    ws.sheet_view.showGridLines = False

    _titulo_hoja(ws, "ANEXO II — VENTA DE AGROQUÍMICOS", "Facturación neta en USD por producto y mes")

    # Orden: Producto | Syngenta | TOTAL USD | Enero..Diciembre
    headers = ["Producto", "Syngenta", "TOTAL USD"] + MESES
    _estilo_header(ws, 4, headers)

    for i, row in enumerate(tabla):
        r = i + 5
        zebra = i % 2 == 0
        es_syngenta = row.get("syngenta", "NO") == "SI"

        # Colorear filas Syngenta diferente
        fill_color = "EBF0F9" if es_syngenta else (COLOR_ZEBRA if zebra else "FFFFFF")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        cell_prod = ws.cell(row=r, column=1, value=row.get("articulo", ""))
        cell_prod.font = Font(name="Calibri", size=9)
        cell_prod.fill = fill

        cell_syn = ws.cell(row=r, column=2, value=row.get("syngenta", "NO"))
        cell_syn.font = Font(name="Calibri", size=9, bold=es_syngenta,
                             color="1A4A8A" if es_syngenta else "333333")
        cell_syn.fill = fill
        cell_syn.alignment = Alignment(horizontal="center")

        total = row.get("Total", 0)
        c_total = ws.cell(row=r, column=3, value=total)
        c_total.number_format = '#,##0.00'
        c_total.font = Font(name="Calibri", size=9, bold=True)
        c_total.fill = fill

        for m_idx, mes in enumerate(MESES):
            val = row.get(mes, 0)
            c = ws.cell(row=r, column=4 + m_idx, value=val if val else None)
            c.number_format = '#,##0.00'
            c.font = Font(name="Calibri", size=9)
            c.fill = fill

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14   # TOTAL USD
    for col in range(4, 16):
        ws.column_dimensions[get_column_letter(col)].width = 13


def _hoja_conciliacion_crm(wb: Workbook, conciliacion: List[Dict]):
    ws = wb.create_sheet("Anexo III — CRM")
    ws.sheet_view.showGridLines = False

    _titulo_hoja(ws, "ANEXO III — CRUCE CRM", "Comparación gestión vs CRM Syngenta")

    headers = [
        "Producto", "Fecha", "Tipo", "Comprobante", "CUIT Cliente",
        "Cant. Gestión", "Cant. CRM", "Δ Cant.",
        "Monto Gestión USD", "Monto CRM USD", "Δ Monto USD",
        "Justificación",
    ]
    _estilo_header(ws, 4, headers)

    for i, row in enumerate(conciliacion):
        r = i + 5
        zebra = i % 2 == 0
        estado = row.get("estado", "OK")

        # Color por estado
        if estado == "DIFERENCIA":
            bg = "FFF3CD"
        elif estado in ("SOLO_GESTION", "SOLO_CRM"):
            bg = "FDECEA"
        else:
            bg = COLOR_ZEBRA if zebra else "FFFFFF"

        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        valores = [
            row.get("producto", ""),
            row.get("fecha", ""),
            row.get("tipo_comprobante", ""),
            row.get("numero_comprobante", ""),
            row.get("cuit_cliente", ""),
            row.get("cantidad_gestion"),
            row.get("cantidad_crm"),
            row.get("diferencia_cantidad"),
            row.get("monto_gestion_usd"),
            row.get("monto_crm_usd"),
            row.get("diferencia_monto"),
            row.get("justificacion", ""),
        ]

        for col_idx, val in enumerate(valores, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.fill = fill
            c.font = Font(name="Calibri", size=9)
            if isinstance(val, float):
                c.number_format = '#,##0.00'
            if col_idx == 12:
                c.alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    for col in ["F", "G", "H", "I", "J", "K"]:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["L"].width = 45
