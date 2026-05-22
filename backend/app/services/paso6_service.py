"""
Paso 6 — Informe con Glosario
Mismo que Paso 5 pero agrega columna de nombre estandarizado del glosario
en Venta de Agroquímicos.
"""
import os
from typing import List, Dict

from ..ai.justificador import mapear_glosario
from .paso5_service import ejecutar_paso5, _generar_excel
from ..core.config import settings


def ejecutar_paso6(
    resumen_compras: List[Dict],
    tabla_apertura: List[Dict],
    conciliacion_crm: List[Dict],
    glosario: List[Dict],
    expediente_info: Dict,
    expediente_id: int,
) -> dict:
    """
    Mismo que Paso 5 + mapeo de glosario en tabla_apertura.
    """
    # Mapear productos al glosario con IA
    productos = [r.get("articulo", "") for r in tabla_apertura]
    mapeos = mapear_glosario(productos, glosario)
    mapa = {m["producto"]: m["nombre_glosario"] for m in mapeos}

    # Agregar columna nombre_glosario a tabla_apertura
    tabla_con_glosario = []
    for row in tabla_apertura:
        new_row = dict(row)
        new_row["nombre_glosario"] = mapa.get(row.get("articulo", ""), row.get("articulo", ""))
        tabla_con_glosario.append(new_row)

    # Filtrar top 90% compras
    from .paso5_service import _filtrar_top90
    resumen_top90 = _filtrar_top90(resumen_compras)

    upload_dir = os.path.join(settings.UPLOAD_DIR, str(expediente_id))
    os.makedirs(upload_dir, exist_ok=True)

    path_excel = os.path.join(upload_dir, "informe_paso6.xlsx")
    _generar_excel_con_glosario(
        resumen_top90, tabla_con_glosario, conciliacion_crm, expediente_info, path_excel
    )

    return {
        "excel_path": path_excel,
        "tabla_con_glosario": tabla_con_glosario,
        "totales": {
            "proveedores_incluidos": len(resumen_top90),
            "lineas_apertura": len(tabla_con_glosario),
            "lineas_crm": len(conciliacion_crm),
        },
    }


def _generar_excel_con_glosario(
    resumen_compras, tabla_apertura, conciliacion_crm, expediente_info, path
):
    """
    Genera el Excel del Paso 6: igual al Paso 5 pero con columna 'Nombre Glosario'
    en el Anexo II.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from .paso5_service import (
        _hoja_portada, _hoja_resumen_compras, _hoja_conciliacion_crm,
        _titulo_hoja, _estilo_header, COLOR_HEADER, COLOR_ZEBRA, MESES
    )

    wb = Workbook()
    wb.remove(wb.active)

    _hoja_portada(wb, expediente_info)
    _hoja_resumen_compras(wb, resumen_compras)
    _hoja_tabla_apertura_con_glosario(wb, tabla_apertura)
    _hoja_conciliacion_crm(wb, conciliacion_crm)

    wb.save(path)


def _hoja_tabla_apertura_con_glosario(wb, tabla: List[Dict]):
    """Tabla de apertura con columna adicional de Nombre Glosario."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from .paso5_service import _titulo_hoja, _estilo_header, COLOR_ZEBRA, MESES

    ws = wb.create_sheet("Anexo II — Agroquímicos")
    ws.sheet_view.showGridLines = False

    _titulo_hoja(ws, "ANEXO II — VENTA DE AGROQUÍMICOS",
                 "Facturación neta en USD por producto y mes — con glosario estandarizado")

    # Headers con columna extra
    headers = ["Producto", "Nombre Glosario", "Syngenta"] + MESES + ["TOTAL USD"]
    _estilo_header(ws, 4, headers)

    for i, row in enumerate(tabla):
        r = i + 5
        zebra = i % 2 == 0
        es_syngenta = row.get("syngenta", "NO") == "SI"
        fill_color = "EBF0F9" if es_syngenta else (COLOR_ZEBRA if zebra else "FFFFFF")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        ws.cell(row=r, column=1, value=row.get("articulo", "")).fill = fill
        glosario_cell = ws.cell(row=r, column=2, value=row.get("nombre_glosario", ""))
        glosario_cell.fill = fill
        glosario_cell.font = Font(name="Calibri", size=9, italic=True, color="1A4A8A")

        syn_cell = ws.cell(row=r, column=3, value=row.get("syngenta", "NO"))
        syn_cell.fill = fill
        syn_cell.alignment = Alignment(horizontal="center")
        syn_cell.font = Font(name="Calibri", size=9, bold=es_syngenta,
                             color="1A4A8A" if es_syngenta else "333333")

        for m_idx, mes in enumerate(MESES):
            val = row.get(mes, 0)
            c = ws.cell(row=r, column=4 + m_idx, value=val if val else None)
            c.number_format = '#,##0.00'
            c.fill = fill
            c.font = Font(name="Calibri", size=9)

        total = row.get("Total", 0)
        c_total = ws.cell(row=r, column=16, value=total)
        c_total.number_format = '#,##0.00'
        c_total.font = Font(name="Calibri", size=9, bold=True)
        c_total.fill = fill

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 10
    for col in range(4, 17):
        ws.column_dimensions[get_column_letter(col)].width = 13
