"""
Paso 3 — Cruce CRM
Compara agroquímicos Syngenta (bajada de gestión) vs CRM de Syngenta.
Genera conciliación con justificaciones IA.
"""
import os
import pandas as pd
import numpy as np
from typing import List, Dict

from ..ai.justificador import generar_justificaciones
from ..ai.smart_parser import parsear_excel
from ..core.config import settings
from .paso1_service import (
    normalizar_tipo_comprobante,
    normalizar_monto,
    normalizar_numero_comprobante,
    es_articulo_no_producto,
)


# Schema CRM Syngenta. ORDEN IMPORTANTE: los campos del DISTRIBUIDOR ("Cuenta:")
# van primero para que el parser "consuma" esas columnas antes que las del
# cliente ("Vendido a:") — así no las confunde.
CRM_SCHEMA = {
    # ── Distribuidor (la "Cuenta" de Syngenta que reporta) ──────────────────
    "cuit_distribuidor": [
        "cuenta: cuit", "cuenta cuit", "cuit cuenta",
        "cuit del distribuidor", "cuit distribuidor", "account cuit",
    ],
    "nombre_distribuidor": [
        "cuenta: nombre", "cuenta nombre", "nombre de la cuenta",
        "nombre cuenta", "razon social cuenta", "razón social cuenta",
        "nombre del distribuidor", "account name",
    ],
    # ── Cliente final del distribuidor ──────────────────────────────────────
    "cuit_cliente_crm": [
        "vendido a: cuit", "vendido a cuit",
        "cuit cliente", "cuit comprador", "cliente cuit",
    ],
    "cliente_crm": [
        "vendido a: nombre", "vendido a nombre",
        "nombre cliente", "razón social cliente",
        "cliente", "razón social", "razon social",
        "denominación", "denominacion", "comprador",
    ],
    # ── Datos del comprobante / línea ───────────────────────────────────────
    "fecha_crm": ["fecha de la factura", "fecha factura", "fecha venta", "fecha", "date"],
    "tipo_crm": ["tipo de documento", "tipo de comprobante", "tipo comprobante", "tipo"],
    "numero_crm": [
        "numero de factura", "número de factura",
        "numero factura", "número factura",
        "nro comprobante", "nro. comprobante",
        "número desde", "numero desde",
        "comprobante", "factura",
    ],
    "producto_crm": [
        "descripción homogénea", "descripcion homogenea",
        "nombre del producto", "producto de lealtad",
        "descripción producto", "descripcion producto",
        "producto", "artículo", "articulo",
        "descripción", "descripcion", "item",
    ],
    # Nombre comercial / marca Syngenta — se usa como GROUND TRUTH para
    # filtrar la gestión: si un producto del distribuidor contiene una marca
    # del CRM, es Syngenta. Más confiable que cualquier clasificación con IA.
    "marca_crm": [
        "nombre comercial", "producto de lealtad: nombre comercial",
        "marca", "brand", "trademark",
    ],
    "cantidad_crm": [
        "volume in normalized", "volumen normalizado", "volumen",
        "cantidad", "cant", "qty", "unidades",
    ],
    "monto_crm": [
        "monto de la factura", "monto factura",
        "monto usd", "importe usd", "total usd",
        "monto", "importe total", "total",
    ],
}

# Exclusiones para evitar colisiones de mapeo entre columnas "Cuenta:" vs
# "Vendido a:" (distribuidor vs cliente final), y para que el número de
# factura no se confunda con un número de cuenta interno de Syngenta.
CRM_EXCLUSIONES = {
    # Para distinguir "Cuenta:" (distribuidor) vs "Vendido a:" (cliente)
    # NO se usan exclusiones por "cuenta" porque "Vendido a: Nombre de la cuenta"
    # también la contiene. La separación viene dada por el ORDEN del schema
    # (distribuidor primero → consume sus columnas) y el used_cols del parser.
    # Las exclusiones acá son solo para evitar choques duros con CUIT/CUIL/PV.
    "cuit_cliente_crm":    ["distribuidor", "numero"],
    "cliente_crm":         ["distribuidor", "cuit", "cuil", "numero"],
    "numero_crm":          ["cuit", "cuil", "cuenta:", "vendido a"],
    "tipo_crm":            ["cuit", "cuil", "cuenta:", "vendido a"],
    "cantidad_crm":        ["monto", "importe", "precio", "price"],
    "producto_crm":        ["codigo", "código", "code"],
}


def _solo_digitos(s) -> str:
    """
    Devuelve solo los dígitos de una cadena. Para comparar CUITs sin formato.

    Maneja el caso openpyxl/read_only en el que un CUIT guardado como número
    en Excel viene como float 30708563311.0 → str sería "30708563311.0" y
    los dígitos darían "307085633110" (12 cifras). Convertir a int primero
    elimina el ".0" espurio.
    """
    if s is None:
        return ""
    if isinstance(s, float):
        s = str(int(s)) if s.is_integer() else str(s)
    return "".join(c for c in str(s) if c.isdigit())


def _norm_header(s) -> str:
    """Normaliza un nombre de columna a minúsculas sin acentos para matchear."""
    import unicodedata
    return "".join(
        c for c in unicodedata.normalize("NFKD", str(s or "").lower())
        if not unicodedata.combining(c)
    ).strip()


def _leer_crm_filtrado_streaming(path: str, cuit_distribuidor: str) -> tuple:
    """
    Lee el archivo CRM filtrando INLINE por el CUIT del distribuidor del
    expediente, en lugar de cargar las 200k+ filas en pandas. Solo quedan en
    memoria las filas del distribuidor actual (~unos miles).

    Usa python-calamine (Rust, ~5x más rápido que openpyxl) si está disponible.
    Cae a openpyxl read-only como fallback. Si ninguno aplica, devuelve
    (None, diag) y el caller usa el camino tradicional con pd.read_excel.
    """
    cuit_norm = _solo_digitos(cuit_distribuidor)
    if not cuit_norm:
        return None, {"motivo": "expediente sin CUIT configurado"}

    # ── Intento 1: calamine (rápido) ────────────────────────────────────────
    try:
        import python_calamine
        wb = python_calamine.CalamineWorkbook.from_path(path)
        sheet = wb.get_sheet_by_index(0)
        data = sheet.to_python()
        if not data:
            return None, {"motivo": "archivo vacío"}
        header = list(data[0])
        rows_iter = iter(data[1:])
        metodo = "streaming (calamine)"
    except Exception as e:
        # ── Intento 2: openpyxl como fallback ───────────────────────────────
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            r_iter = ws.iter_rows(values_only=True)
            try:
                header = list(next(r_iter))
            except StopIteration:
                wb.close()
                return None, {"motivo": "archivo vacío"}
            rows_iter = r_iter
            metodo = "streaming (openpyxl)"
            # cerramos wb al final, no lo cerramos acá porque iter es perezoso
        except Exception as e2:
            return None, {"motivo": f"no se pudo leer como xlsx (calamine: {e}; openpyxl: {e2})"}

    # ── Localizar columnas de CUIT y nombre del distribuidor ───────────────
    cuit_col_idx = None
    nombre_col_idx = None
    for i, col in enumerate(header):
        cn = _norm_header(col)
        if cuit_col_idx is None and (
            ("cuenta" in cn and "cuit" in cn) or
            ("distribuidor" in cn and "cuit" in cn)
        ):
            cuit_col_idx = i
        if nombre_col_idx is None and (
            ("cuenta" in cn and ("nombre" in cn or "razon" in cn)) or
            ("nombre" in cn and "distribuidor" in cn)
        ):
            nombre_col_idx = i
    if cuit_col_idx is None:
        return None, {"motivo": "no se encontró columna Cuenta: CUIT en el header"}

    # ── Iterar filas, filtrar + enumerar distribuidores en una pasada ──────
    filtradas = []
    total = 0
    distribuidores: dict = {}
    for row in rows_iter:
        total += 1
        if cuit_col_idx >= len(row):
            continue
        val = row[cuit_col_idx]
        if val is None:
            continue
        this_cuit = _solo_digitos(val)
        if not this_cuit:
            continue
        if this_cuit not in distribuidores:
            nombre = ""
            if nombre_col_idx is not None and nombre_col_idx < len(row):
                nombre = str(row[nombre_col_idx] or "")
            distribuidores[this_cuit] = [nombre, 0]
        distribuidores[this_cuit][1] += 1
        if this_cuit == cuit_norm:
            filtradas.append(row)

    diag = {
        "filas_archivo_total": total,
        "filas_distribuidor": len(filtradas),
        "metodo_lectura": metodo,
        "distribuidores_archivo": sorted(
            [{"cuit": c, "nombre": n, "filas": f} for c, (n, f) in distribuidores.items()],
            key=lambda x: -x["filas"],
        ),
    }

    if not filtradas:
        return pd.DataFrame(columns=header), diag

    df = pd.DataFrame(filtradas, columns=header)
    return df, diag


def _extraer_marcas_syngenta(df_crm: pd.DataFrame) -> set:
    """
    Devuelve el conjunto de marcas Syngenta presentes en el CRM filtrado al
    distribuidor. Estas marcas son el GROUND TRUTH para identificar productos
    Syngenta en la gestión del distribuidor — más confiable que una
    clasificación con IA.

    Prioridad:
      1. Columna "marca_crm" (Nombre comercial — la marca pura, ej "MEGAFOL")
      2. Primera palabra de "producto_crm" como proxy (ej "MIRAVIS" de
         "MIRAVIS DUO 4X5 L")
    """
    marcas = set()
    for col in ("marca_crm", "producto_crm"):
        if col not in df_crm.columns:
            continue
        for v in df_crm[col].dropna().unique():
            s = str(v).strip().upper()
            if not s or s == "NAN":
                continue
            if col == "marca_crm":
                marcas.add(s)
            else:
                # Primera palabra significativa (>2 chars) como marca proxy
                primera = s.split()[0] if s.split() else ""
                if len(primera) > 2 and primera.isalpha():
                    marcas.add(primera)
    # Quitar palabras genéricas que generarían falsos positivos
    GENERICAS = {"PRODUCTO", "ARTICULO", "ITEM", "DESC", "VARIOS", "OTROS"}
    return marcas - GENERICAS


def _filtrar_gestion_por_marcas_crm(df_gestion: pd.DataFrame, marcas: set) -> pd.DataFrame:
    """
    Filtra la gestión al subset Syngenta usando las marcas del CRM como
    referencia. Más preciso que la clasificación con IA porque el CRM ES la
    fuente oficial de Syngenta.

    Incluye:
      - Cualquier línea cuyo articulo contenga ALGUNA marca del CRM
      - Todas las NC/ND (los archivos de NC/ND no traen articulo y pueden
        afectar productos Syngenta — quedarán como SOLO_GESTION si el CRM
        no las tiene, lo cual es una señal honesta para el auditor)
    """
    if df_gestion.empty:
        return df_gestion
    df = df_gestion.copy()
    if "articulo" not in df.columns:
        df["articulo"] = ""
    if "tipo" not in df.columns:
        df["tipo"] = ""

    art_upper = df["articulo"].fillna("").astype(str).str.upper()
    es_nc_nd = df["tipo"].astype(str).isin(["NC", "ND"])

    # Excluir líneas que son ajustes financieros (AJUSTE DE PRECIO, INTERES,
    # RECARGO, etc.) — no son productos vendidos, distorsionan el cruce con
    # CRM Syngenta y el informe del Paso 5.
    es_no_prod = df["articulo"].apply(es_articulo_no_producto)
    n_no_prod = int(es_no_prod.sum())
    if n_no_prod:
        print(f"[PASO 3] Excluidas {n_no_prod} líneas que son ajustes financieros "
              f"(AJUSTE/INTERES/RECARGO/etc) — no son productos")

    if not marcas:
        # Sin info de marcas → quedarse al menos con NC/ND (sin ajustes)
        return df[es_nc_nd & ~es_no_prod].copy()

    # Match si CUALQUIER marca está contenida en el articulo
    matches = art_upper.apply(lambda s: any(b in s for b in marcas))
    return df[(matches | es_nc_nd) & ~es_no_prod].copy()


def _numero_crm_a_estandar(valor: str) -> str:
    """
    El CRM de Syngenta usa un formato compuesto en 'Numero de factura':
        'A002226061|FC|A000600020954'
        (cuenta) | (tipo) | (numero fiscal compuesto)

    Esta función extrae la parte fiscal y la normaliza al mismo formato que
    usa Paso 1 ("PPPPP-NNNNNNNN") para que el cruce matchee.
    """
    if not valor or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    s = str(valor).strip()
    # Si tiene '|', tomar la última parte (el número fiscal compuesto)
    if "|" in s:
        s = s.split("|")[-1]
    # Extraer solo dígitos y normalizar usando la misma función de Paso 1
    digitos = _solo_digitos(s)
    if not digitos:
        return ""
    return normalizar_numero_comprobante(valor_combinado=digitos)


def leer_crm(path: str):
    """Lee reporte CRM usando smart_parser. Retorna (df_renombrado, info_parser)."""
    resultado = parsear_excel(
        path=path,
        task_id="crm",
        schema=CRM_SCHEMA,
        excluir_si_contiene=CRM_EXCLUSIONES,
        # cuit_distribuidor es crítico — sin él no podemos filtrar al distribuidor
        # del expediente cuando el CRM contiene reportes de varios distribuidores.
        columnas_criticas=["fecha_crm", "numero_crm", "producto_crm", "monto_crm",
                          "cuit_distribuidor"],
    )
    df = resultado["df"]
    mapping = resultado["mapping"]

    rename_map = {orig: std for std, orig in mapping.items() if orig}
    df = df.rename(columns=rename_map)

    print(f"[CRM] método={resultado['metodo']} conf={resultado['confianza']:.2f} "
          f"mapping={mapping} warnings={resultado['warnings']}")
    return df, resultado


def _filtrar_crm_por_distribuidor(df: pd.DataFrame, cuit_exp: str,
                                  nombre_exp: str = "") -> tuple:
    """
    Filtra el CRM dejando SOLO las filas del distribuidor del expediente.

    Estrategia:
      1. Match exacto por CUIT (dígitos puros) — determinístico, sin IA, infalible.
      2. Si no hay match por CUIT, intenta fuzzy match por nombre (substring
         insensible a mayúsculas / acentos) como fallback de seguridad.
      3. Si tampoco así, levanta un error claro listando los distribuidores
         que SÍ están en el archivo para que el auditor verifique.

    Devuelve (df_filtrado, diag) con info de qué pasó.
    """
    diag = {
        "filas_archivo_total": int(len(df)),
        "filas_distribuidor": 0,
        "metodo_filtro": None,
        "distribuidores_archivo": [],   # primeros N para el mensaje de error
    }
    if "cuit_distribuidor" not in df.columns:
        raise ValueError(
            "El archivo CRM no tiene una columna que identifique al distribuidor "
            "(se esperaba 'Cuenta: CUIT' o similar). Sin esa columna no se puede "
            "filtrar al distribuidor del expediente del CRM consolidado."
        )

    cuit_norm = _solo_digitos(cuit_exp or "")

    # ── 1. Match exacto por CUIT ─────────────────────────────────────────────
    df = df.copy()
    df["__cuit_dist_norm__"] = df["cuit_distribuidor"].astype(str).apply(_solo_digitos)

    if cuit_norm:
        sel = df[df["__cuit_dist_norm__"] == cuit_norm]
        if len(sel) > 0:
            diag["filas_distribuidor"] = int(len(sel))
            diag["metodo_filtro"] = f"CUIT exacto ({cuit_exp})"
            return sel.drop(columns=["__cuit_dist_norm__"]), diag

    # ── 2. Fallback fuzzy por nombre ─────────────────────────────────────────
    if nombre_exp and "nombre_distribuidor" in df.columns:
        import unicodedata
        def _norm(s):
            return "".join(c for c in unicodedata.normalize("NFKD", str(s).lower())
                          if not unicodedata.combining(c)).strip()
        nombre_norm = _norm(nombre_exp)
        if nombre_norm:
            df["__nom_norm__"] = df["nombre_distribuidor"].astype(str).apply(_norm)
            # Match si el nombre del expediente está contenido en el nombre del CRM
            # o viceversa (atajamos diferencias tipo "X SA" vs "X").
            sel = df[df["__nom_norm__"].str.contains(nombre_norm, na=False, regex=False)]
            if len(sel) == 0:
                # Intentar al revés: nombre del CRM contenido en el del expediente
                sel = df[df["__nom_norm__"].apply(
                    lambda v: bool(v) and v in nombre_norm
                )]
            if len(sel) > 0:
                diag["filas_distribuidor"] = int(len(sel))
                diag["metodo_filtro"] = f"nombre similar ({nombre_exp})"
                return sel.drop(columns=["__cuit_dist_norm__", "__nom_norm__"]), diag

    # ── 3. Ningún match: levantar error con la lista de distribuidores ──────
    distrib_disponibles = (
        df.groupby(["__cuit_dist_norm__", "nombre_distribuidor"], dropna=False)
          .size().reset_index(name="filas")
          .sort_values("filas", ascending=False)
          .head(15)
    ) if "nombre_distribuidor" in df.columns else (
        df.groupby("__cuit_dist_norm__").size().reset_index(name="filas")
        .sort_values("filas", ascending=False).head(15)
    )

    ejemplos = []
    for _, r in distrib_disponibles.iterrows():
        nombre = r.get("nombre_distribuidor", "(s/n)") if "nombre_distribuidor" in r.index else ""
        ejemplos.append(f"  - CUIT {r['__cuit_dist_norm__']}: {nombre} ({int(r['filas'])} filas)")
    diag["distribuidores_archivo"] = ejemplos

    raise ValueError(
        f"En el archivo CRM no hay datos del distribuidor con CUIT '{cuit_exp}' "
        f"({nombre_exp or 'sin nombre'}). El archivo contiene "
        f"{df['__cuit_dist_norm__'].nunique()} distribuidores distintos. "
        f"Los principales son:\n" + "\n".join(ejemplos[:10]) +
        "\n\nVerificá que el CUIT del expediente esté bien escrito y que el CRM "
        "incluya el período del distribuidor que estás auditando."
    )


def _find_col(cols_lower: dict, keywords: list):
    for kw in keywords:
        for col_l, col_orig in cols_lower.items():
            if kw in col_l:
                return col_orig
    return None


def ejecutar_paso3(
    path_agroquimicos_syngenta: str,
    path_crm: str,
    expediente_id: int,
    cuit_distribuidor: str = "",
    nombre_distribuidor: str = "",
    anotaciones_manuales: list = None,
) -> dict:
    """
    Cruza agroquímicos Syngenta (Paso 2) vs CRM.

    El archivo CRM de Syngenta contiene reportes de TODOS los distribuidores;
    se filtra al distribuidor del expediente por CUIT antes de cruzar.

    anotaciones_manuales: list de dicts con anotaciones manuales del Paso 1
    (comprobantes ARCA que el auditor completó a mano porque la gestión
    auto-detectada no los tenía). Se agregan al lado gestión del cruce para
    que aparezcan como matches en lugar de SOLO_CRM erróneos.
    """
    # Cargar datos
    df_agro = pd.read_excel(path_agroquimicos_syngenta, dtype=str)

    # ── Lectura del CRM con FILTRO INLINE (streaming) ───────────────────────
    # El CRM puede tener 200k+ filas (todos los distribuidores). Leer todo
    # con pd.read_excel y filtrar después tumba el worker por memoria/tiempo.
    # Estrategia: openpyxl read_only + iter_rows, descartando inline las filas
    # que no son del distribuidor del expediente. Solo cargamos en memoria las
    # filas relevantes (~6k de 188k en el caso real).
    df_crm_stream, stream_diag = _leer_crm_filtrado_streaming(path_crm, cuit_distribuidor)

    # Si streaming corrió OK pero no encontró filas para este CUIT, levantar
    # error claro con la lista de distribuidores presentes (todo desde la
    # MISMA pasada de streaming, sin volver a leer el archivo).
    if df_crm_stream is not None and df_crm_stream.empty:
        ejemplos = stream_diag.get("distribuidores_archivo", [])[:10]
        lista = "\n".join(
            f"  - CUIT {d['cuit']}: {d['nombre'] or '(sin nombre)'} ({d['filas']:,} filas)"
            for d in ejemplos
        )
        raise ValueError(
            f"En el archivo CRM no hay datos del distribuidor con CUIT "
            f"'{cuit_distribuidor}' ({nombre_distribuidor or 'sin nombre'}). "
            f"El archivo tiene {len(stream_diag.get('distribuidores_archivo', []))} "
            f"distribuidores distintos en {stream_diag.get('filas_archivo_total', 0):,} "
            f"filas. Los principales son:\n" + lista +
            "\n\nVerificá que el CUIT del expediente esté bien escrito (editalo "
            "con el botón ✏️ del Dashboard) y que el CRM incluya el período del "
            "distribuidor que estás auditando."
        )

    if df_crm_stream is not None and not df_crm_stream.empty:
        # Streaming OK y encontró filas → seguir con el df filtrado.
        # Aplicamos detección de columnas con keywords sobre el df ya pequeño.
        from ..ai.smart_parser import detectar_por_keywords
        mapping = detectar_por_keywords(
            [str(c) for c in df_crm_stream.columns], CRM_SCHEMA, CRM_EXCLUSIONES,
        )
        rename_map = {orig: std for std, orig in mapping.items() if orig}
        df_crm = df_crm_stream.rename(columns=rename_map)
        crm_info = {
            "metodo": "streaming",
            "confianza": 1.0,
            "mapping": mapping,
            "columnas_archivo": [str(c) for c in df_crm_stream.columns],
            "columnas_faltantes": [c for c in CRM_SCHEMA if not mapping.get(c)],
            "columnas_no_mapeadas": [c for c in df_crm_stream.columns if c not in mapping.values()],
            "warnings": [],
        }
        filtro_diag = {
            "filas_archivo_total": stream_diag["filas_archivo_total"],
            "filas_distribuidor": stream_diag["filas_distribuidor"],
            "metodo_filtro": f"CUIT exacto streaming ({cuit_distribuidor})",
            "distribuidores_archivo": [],
        }
        print(f"[CRM] Streaming → {filtro_diag['filas_distribuidor']:,} de "
              f"{filtro_diag['filas_archivo_total']:,} filas (sin cargar el resto en memoria)")
    else:
        # Fallback: streaming no aplica (no es xlsx, header sin 'Cuenta: CUIT',
        # etc.) — caer al camino tradicional que carga todo y filtra después.
        # Este camino también es el que dispara el mensaje de error útil
        # listando los distribuidores presentes si el CUIT no matchea.
        motivo = (stream_diag or {}).get("motivo", "desconocido")
        print(f"[CRM] Streaming no aplicable ({motivo}) — fallback a lectura completa")
        df_crm_full, crm_info = leer_crm(path_crm)
        df_crm, filtro_diag = _filtrar_crm_por_distribuidor(
            df_crm_full, cuit_distribuidor, nombre_distribuidor,
        )
        print(f"[CRM] Filtro distribuidor: {filtro_diag['metodo_filtro']} → "
              f"{filtro_diag['filas_distribuidor']:,} de "
              f"{filtro_diag['filas_archivo_total']:,} filas")

    # Preparar DataFrame de gestión (puede ser bajada_normalizada completa o
    # bajada_syngenta ya filtrada — _preparar_gestion tolera ambos esquemas).
    df_gestion = _preparar_gestion(df_agro)

    # ── Anotaciones manuales del Paso 1 ─────────────────────────────────────
    # Solo se agregan al cruce las que vienen COMPLETAS:
    #   - monto_gestion_usd > 0 (sin monto no aportan al match)
    #   - producto con texto (sino no podemos cruzar con CRM por producto)
    # Las anotaciones a medio llenar (cliente sin monto, sin clasificación) se
    # ignoran — meterlas al cruce solo agregaría ruido.
    if anotaciones_manuales:
        completas = [
            a for a in anotaciones_manuales
            if (a.get("monto_usd") or 0) != 0 and str(a.get("articulo") or "").strip()
        ]
        if completas:
            df_an = _preparar_gestion(pd.DataFrame(completas))
            df_gestion = pd.concat([df_gestion, df_an], ignore_index=True)
            print(f"[PASO 3] +{len(df_an)} anotaciones manuales completas agregadas "
                  f"(de {len(anotaciones_manuales)} totales)")

    df_crm = _preparar_crm(df_crm)

    # ── Filtrar gestión usando el CRM como GROUND TRUTH ──────────────────────
    # En vez de confiar en la clasificación con IA del Paso 2 (que tiene falsos
    # positivos y negativos), usamos las marcas que aparecen en el CRM del
    # distribuidor para decidir qué líneas de gestión son Syngenta. Esto
    # garantiza que las dos sumas que comparamos (gestión vs CRM) midan
    # exactamente el mismo universo de productos.
    marcas_crm = _extraer_marcas_syngenta(df_crm)
    if marcas_crm:
        n_antes = len(df_gestion)
        df_gestion = _filtrar_gestion_por_marcas_crm(df_gestion, marcas_crm)
        print(f"[PASO 3] Filtro gestión por marcas Syngenta del CRM "
              f"({len(marcas_crm)} marcas): {len(df_gestion):,} de {n_antes:,} líneas")

    # Cruce
    conciliacion = _cruzar_crm(df_gestion, df_crm)

    # Generar justificaciones con IA SOLO para las verdaderas DIFERENCIAS
    # (matchearon en ambos lados pero los montos no cuadran). Las SOLO_GESTION
    # y SOLO_CRM ya tienen una justificación template ("Venta sin reporte en
    # CRM" / "Reportado en CRM sin factura") — la IA no aporta nada ahí y son
    # la MAYORÍA de las filas (lo que disparaba los $13 de tokens).
    #
    # Además se ordenan por monto descendente y se capean a 200 — más que eso
    # no aporta valor y empieza a costar caro / tardar minutos. Las que sobran
    # quedan como pendientes para análisis manual.
    MAX_IA = 200
    diferencias_reales = [
        r for r in conciliacion
        if r.get("estado") == "DIFERENCIA"
        and abs(r.get("diferencia_monto") or 0) > 0.01
    ]
    diferencias_reales.sort(key=lambda r: abs(r.get("diferencia_monto") or 0), reverse=True)
    overflow = []
    if len(diferencias_reales) > MAX_IA:
        overflow = diferencias_reales[MAX_IA:]
        diferencias_reales = diferencias_reales[:MAX_IA]
        for r in overflow:
            r["justificacion"] = (
                f"Pendiente — más de {MAX_IA} diferencias para analizar con IA. "
                f"Revisar manualmente."
            )

    if diferencias_reales:
        try:
            justificaciones = generar_justificaciones(diferencias_reales)
        except Exception as e:
            print(f"[PASO 3] generar_justificaciones falló: {e}")
            justificaciones = []
        for r, j in zip(diferencias_reales, justificaciones):
            if j:
                r["justificacion"] = j

    print(f"[PASO 3] IA: {len(diferencias_reales)} diferencias procesadas"
          f"{' (+' + str(len(overflow)) + ' overflow)' if overflow else ''}, "
          f"SOLO_* con template (sin IA)")

    resumen = {
        "total_lineas": len(conciliacion),
        "sin_diferencia": sum(1 for r in conciliacion if r.get("estado") == "OK"),
        "con_diferencia": sum(1 for r in conciliacion if r.get("estado") == "DIFERENCIA"),
        "solo_gestion": sum(1 for r in conciliacion if r.get("estado") == "SOLO_GESTION"),
        "solo_crm": sum(1 for r in conciliacion if r.get("estado") == "SOLO_CRM"),
    }

    # Sumar el diagnóstico del filtro de distribuidor a los warnings del parser
    warnings_crm = list(crm_info.get("warnings", []))
    warnings_crm.insert(0,
        f"Filtro distribuidor: {filtro_diag['metodo_filtro']} — "
        f"{filtro_diag['filas_distribuidor']:,} de "
        f"{filtro_diag['filas_archivo_total']:,} filas del archivo CRM"
    )

    # Diagnóstico estructurado
    parser_diagnostico = []
    if crm_info:
        parser_diagnostico.append({
            "archivo": "CRM Syngenta",
            "metodo": crm_info["metodo"],
            "confianza": crm_info["confianza"],
            "mapping": crm_info["mapping"],
            "columnas_archivo": crm_info.get("columnas_archivo", []),
            "columnas_faltantes": crm_info.get("columnas_faltantes", []),
            "columnas_no_mapeadas": crm_info.get("columnas_no_mapeadas", []),
            "warnings": warnings_crm,
        })

    return {
        "conciliacion": conciliacion,
        "resumen": resumen,
        "parser_diagnostico": parser_diagnostico,
        "filtro_distribuidor": filtro_diag,
    }


def _preparar_gestion(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normaliza el DataFrame de gestión Syngenta para el cruce con CRM.

    Es tolerante a los dos esquemas en los que puede venir el archivo:
      - Esquema nuevo (Paso 1 line-item): tipo, numero, fecha, cliente,
        cuit_cliente, articulo, monto_usd, ...
      - Esquema viejo: tipo_comprobante, numero_comprobante, ...
    Y también a columnas faltantes (ej: 'cantidad' puede no existir todavía).
    """
    df = df.copy()
    cols = set(df.columns)

    def _col_or_default(*names, default=""):
        """Devuelve la primera columna que exista como Series, sino una Serie con default."""
        for n in names:
            if n in cols:
                return df[n]
        return pd.Series([default] * len(df), index=df.index)

    df["fecha"] = pd.to_datetime(_col_or_default("fecha", default=""), errors="coerce")

    # tipo_comprobante: nuevo "tipo" o viejo "tipo_comprobante"
    tipo_series = _col_or_default("tipo_comprobante", "tipo", default="FC")
    df["tipo_comprobante"] = tipo_series.astype(str).apply(normalizar_tipo_comprobante)

    df["numero_comprobante"] = _col_or_default(
        "numero_comprobante", "numero", default=""
    ).fillna("").astype(str).str.strip()
    # cuit normalizado a dígitos puros para que matchee con el cuit del CRM
    df["cuit_cliente"] = _col_or_default("cuit_cliente", default="").fillna("").astype(str).apply(_solo_digitos)
    # IMPORTANTE: fillna("") ANTES de astype(str). Sino NaN → "nan" → "NAN"
    # tras upper(), y aparece "NAN" como producto en filas NC/ND sin detalle.
    df["articulo"] = _col_or_default("articulo", default="").fillna("").astype(str).str.strip().str.upper()

    # cantidad puede no existir si la bajada se generó con la versión vieja
    # de Paso 1 (que no la capturaba) — default 0.
    df["cantidad"] = pd.to_numeric(
        _col_or_default("cantidad", default=0), errors="coerce"
    ).fillna(0)

    df["monto_usd"] = pd.to_numeric(
        _col_or_default("monto_usd", "monto_total", default=0), errors="coerce"
    ).fillna(0)
    return df


def _preparar_crm(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    cols = set(df.columns)

    def _serie(*names, default=""):
        for n in names:
            if n in cols:
                return df[n]
        return pd.Series([default] * len(df), index=df.index)

    df["fecha_crm"] = pd.to_datetime(_serie("fecha_crm", default=""), errors="coerce")
    # Normalizar el "Numero de factura" del CRM (formato "AXXX|FC|A000600020954")
    # al mismo formato que usa Paso 1 (PPPPP-NNNNNNNN), sino el cruce nunca matchea.
    df["numero_crm"] = _serie("numero_crm", default="").astype(str).apply(_numero_crm_a_estandar)
    df["cuit_cliente_crm"] = _serie("cuit_cliente_crm", default="").astype(str).apply(_solo_digitos)
    df["producto_crm"] = _serie("producto_crm", default="").astype(str).str.strip().str.upper()
    df["cliente_crm"] = _serie("cliente_crm", default="").astype(str).str.strip()
    df["cantidad_crm"] = pd.to_numeric(_serie("cantidad_crm", default=0), errors="coerce").fillna(0)
    df["monto_crm"] = pd.to_numeric(_serie("monto_crm", default=0), errors="coerce").fillna(0)
    # tipo_crm: el CRM Syngenta usa "FC"/"NC"/"RM" en "Tipo de documento".
    df["tipo_crm"] = _serie("tipo_crm", default="").astype(str).apply(normalizar_tipo_comprobante)

    # Excluir REMITOS — son entregas de mercadería de Syngenta al distribuidor,
    # NO facturas que él emite. El distribuidor nunca va a tener estos números
    # en su sistema de facturación → matchearlos genera 100+ SOLO_CRM falsos.
    n_total = len(df)
    df = df[df["tipo_crm"].isin(["FC", "NC", "ND"])].copy()
    descartados = n_total - len(df)
    if descartados:
        print(f"[CRM] Excluidos {descartados} comprobantes que no son facturas "
              f"(remitos / movimientos internos)")
    return df


def _cruzar_crm(df_gestion: pd.DataFrame, df_crm: pd.DataFrame) -> List[Dict]:
    """
    Cruce gestión vs CRM AGREGADO POR COMPROBANTE.

    Por qué agregado y no por línea:
      - Los archivos de NC/ND en gestión no traen artículo ni CUIT del cliente
        (son a nivel comprobante), por eso un match línea-a-línea
        articulo+numero+cuit nunca encontraba nada para las NC.
      - Las FACTURAS del ERP a veces no traen Punto de Venta (PV=00000),
        mientras que el CRM sí tiene el PV real → los números no coincidirían
        exactamente aunque sean el mismo comprobante.
      - Los nombres de producto difieren entre el ERP del distribuidor y la
        "Descripción Homogénea" de Syngenta — matching exacto falla.

    Solución: agrupar ambos lados por NÚMERO FISCAL (los últimos 8 dígitos del
    número, ignorando PV y prefijos). Sumar cantidad y monto por comprobante y
    comparar. Esto es lo que un auditor mira: "¿cuánto facturó el distribuidor
    por este comprobante vs cuánto reportó al CRM?".

    Cada fila de salida = un comprobante. Productos quedan como texto
    informativo (lista comma-separated de lo que aparece en cada lado).
    """
    def _num_fiscal(s):
        """Últimos 8 dígitos del número de comprobante (la parte fiscal)."""
        d = "".join(c for c in str(s or "") if c.isdigit())
        return d[-8:].zfill(8) if d else ""

    def _tipo_norm(t):
        """Normaliza el tipo a FC/NC/ND para que matche entre gestión y CRM.
        Importante usar el tipo en la clave de match para no confundir FC nro
        X con NC nro X (mismo número fiscal, comprobantes distintos).
        Los RM (remitos) deben venir ya filtrados de _preparar_crm porque no
        son facturas — si aparece alguno acá se trata como su propio tipo
        para que NO matchee con FCs del distribuidor."""
        t = str(t or "").upper().strip()
        if "NC" in t or t.startswith("N/C"):
            return "NC"
        if "ND" in t or t.startswith("N/D"):
            return "ND"
        if t == "RM" or "REMITO" in t:
            return "RM"
        return "FC"   # FC, FB y resto quedan como FC

    def _primer_no_vacio(serie):
        for v in serie:
            sv = str(v or "").strip()
            if sv:
                return sv
        return ""

    def _productos_concat(serie, max_len=300):
        # Filtrar valores vacíos, NaN literal "nan"/"NAN" y "NONE" — son ruido
        # que aparecía como producto en filas NC/ND sin detalle.
        BASURA = {"", "NAN", "NONE", "NULL"}
        unicos = sorted({
            str(v).strip() for v in serie
            if str(v or "").strip() and str(v).strip().upper() not in BASURA
        })
        s = ", ".join(unicos)
        return s if len(s) <= max_len else s[:max_len - 1] + "…"

    # ── Preparar gestión ────────────────────────────────────────────────────
    g = df_gestion.copy()
    g["__num__"] = g["numero_comprobante"].apply(_num_fiscal)
    if "tipo_comprobante" not in g.columns:
        g["tipo_comprobante"] = ""
    g["__tipo__"] = g["tipo_comprobante"].apply(_tipo_norm)
    g["__key__"] = g["__tipo__"] + "|" + g["__num__"]
    g = g[g["__num__"] != ""]
    if "articulo" not in g.columns:
        g["articulo"] = ""

    g_agg = g.groupby("__key__", dropna=False).agg(
        numero=("numero_comprobante", _primer_no_vacio),
        fecha=("fecha", "first"),
        cuit=("cuit_cliente", _primer_no_vacio),
        tipo=("__tipo__", "first"),
        productos=("articulo", _productos_concat),
        cantidad=("cantidad", "sum"),
        monto=("monto_usd", "sum"),
    ).reset_index()

    # ── Preparar CRM ────────────────────────────────────────────────────────
    c = df_crm.copy()
    c["__num__"] = c["numero_crm"].apply(_num_fiscal)
    if "tipo_crm" not in c.columns:
        c["tipo_crm"] = ""
    c["__tipo__"] = c["tipo_crm"].apply(_tipo_norm)
    c["__key__"] = c["__tipo__"] + "|" + c["__num__"]
    c = c[c["__num__"] != ""]
    if "cliente_crm" not in c.columns:
        c["cliente_crm"] = ""

    c_agg = c.groupby("__key__", dropna=False).agg(
        numero=("numero_crm", _primer_no_vacio),
        fecha=("fecha_crm", "first"),
        cuit=("cuit_cliente_crm", _primer_no_vacio),
        cliente=("cliente_crm", _primer_no_vacio),
        tipo=("__tipo__", "first"),
        productos=("producto_crm", _productos_concat),
        cantidad=("cantidad_crm", "sum"),
        monto=("monto_crm", "sum"),
    ).reset_index()

    g_map = {r["__key__"]: r for _, r in g_agg.iterrows()}
    c_map = {r["__key__"]: r for _, r in c_agg.iterrows()}

    def _fmt_fecha(v):
        s = str(v or "")[:10]
        return s if s and s != "NaT" else ""

    conciliacion = []
    for num in set(g_map) | set(c_map):
        gr = g_map.get(num)
        cr = c_map.get(num)

        if gr is not None and cr is not None:
            cant_g = float(gr["cantidad"] or 0)
            cant_c = float(cr["cantidad"] or 0)
            monto_g = float(gr["monto"] or 0)
            monto_c = float(cr["monto"] or 0)
            diff_monto = round(monto_g - monto_c, 2)

            # Detalle de producto en gestión: si NO tiene (típico de NC/ND
            # cuyo archivo no trae columna de producto), no se puede comparar
            # el monto Syngenta-específico contra el CRM — el monto del lado
            # gestión es el TOTAL del comprobante (todos los productos del
            # cliente), no solo la porción Syngenta. Hacer comparación parcial:
            # OK por presencia.
            gestion_tiene_detalle = bool(str(gr["productos"] or "").strip())

            # CUIT distinto entre gestión y CRM es un finding genuino aun
            # cuando los montos coincidan (lo detecta el auditor).
            cuit_g_d = _solo_digitos(gr["cuit"])
            cuit_c_d = _solo_digitos(cr["cuit"])
            cuit_difiere = bool(cuit_g_d and cuit_c_d and cuit_g_d != cuit_c_d)

            # Tolerancia 5% relativo o $5 absoluto — calibrada contra los
            # umbrales que usan auditores humanos. Diferencias chicas son ruido
            # del TC distinto entre Syngenta y el TC del usuario.
            ABS_TOL = 5.0
            REL_TOL = 0.05
            max_monto = max(abs(monto_g), abs(monto_c))
            rel = abs(diff_monto) / max_monto if max_monto > 0 else 0

            justif = ""
            if cuit_difiere:
                estado = "DIFERENCIA"
                justif = f"CUIT difiere — gestión: {cuit_g_d} vs CRM: {cuit_c_d}"
            elif not gestion_tiene_detalle and abs(monto_g) > abs(monto_c) * 1.10:
                # NC/ND sin detalle de producto Y monto gestión >> monto CRM:
                # típico de NC que cubre varios productos del cliente, donde el
                # CRM solo refleja la porción Syngenta. No se puede comparar
                # montos justamente → OK por presencia.
                estado = "OK"
                justif = ("Reportada en CRM (NC sin detalle de producto en "
                          "gestión — el monto del lado gestión incluye otros "
                          "productos no Syngenta del mismo cliente)")
            else:
                # Comparación normal (vale para FC con detalle y para NC
                # Syngenta-only donde gestion ≈ CRM).
                estado = "OK" if abs(diff_monto) <= ABS_TOL or rel <= REL_TOL else "DIFERENCIA"
                if estado == "DIFERENCIA":
                    justif = "Pendiente de análisis"

            tipo = str(gr["tipo"] or cr["tipo"] or "")
            if not tipo:
                tipo = "NC" if (monto_c < 0 or monto_g < 0) else "FC"
            conciliacion.append({
                "numero_comprobante": str(gr["numero"] or cr["numero"]),
                "fecha": _fmt_fecha(gr["fecha"] or cr["fecha"]),
                "tipo_comprobante": tipo,
                "cuit_cliente": gr["cuit"] or cr["cuit"],
                "cliente": cr["cliente"],
                "producto": gr["productos"] or cr["productos"],
                "cantidad_gestion": round(cant_g, 4),
                "cantidad_crm": round(cant_c, 4),
                "diferencia_cantidad": round(cant_g - cant_c, 4),
                "monto_gestion_usd": round(monto_g, 2),
                "monto_crm_usd": round(monto_c, 2),
                "diferencia_monto": diff_monto,
                "justificacion": justif,
                "estado": estado,
            })
        elif gr is not None:
            monto_g = float(gr["monto"] or 0)
            tipo = str(gr["tipo"]) or ("NC" if monto_g < 0 else "FC")
            conciliacion.append({
                "numero_comprobante": str(gr["numero"]),
                "fecha": _fmt_fecha(gr["fecha"]),
                "tipo_comprobante": tipo,
                "cuit_cliente": gr["cuit"],
                "cliente": "",
                "producto": gr["productos"],
                "cantidad_gestion": round(float(gr["cantidad"] or 0), 4),
                "cantidad_crm": 0,
                "diferencia_cantidad": round(float(gr["cantidad"] or 0), 4),
                "monto_gestion_usd": round(monto_g, 2),
                "monto_crm_usd": 0,
                "diferencia_monto": round(monto_g, 2),
                "justificacion": "Venta facturada sin reporte en CRM",
                "estado": "SOLO_GESTION",
            })
        else:
            monto_c = float(cr["monto"] or 0)
            tipo = str(cr["tipo"]) or ("NC" if monto_c < 0 else "FC")
            conciliacion.append({
                "numero_comprobante": str(cr["numero"]),
                "fecha": _fmt_fecha(cr["fecha"]),
                "tipo_comprobante": tipo,
                "cuit_cliente": cr["cuit"],
                "cliente": cr["cliente"],
                "producto": cr["productos"],
                "cantidad_gestion": 0,
                "cantidad_crm": round(float(cr["cantidad"] or 0), 4),
                "diferencia_cantidad": -round(float(cr["cantidad"] or 0), 4),
                "monto_gestion_usd": 0,
                "monto_crm_usd": round(monto_c, 2),
                "diferencia_monto": -round(monto_c, 2),
                "justificacion": "Reportado en CRM sin factura correspondiente en gestión",
                "estado": "SOLO_CRM",
            })

    # Limpiar ruido: NC/ND del distribuidor que (a) no tienen detalle de
    # producto y (b) no matchearon con nada en CRM. Sin detalle no podemos
    # confirmar si eran Syngenta, y como tampoco aparecen en el CRM Syngenta,
    # casi seguro son NC/ND para productos NO Syngenta. Mostrarlas solo
    # generaría ruido en la conciliación (cientos de filas sin valor para el
    # auditor humano, que las descarta manualmente).
    #
    # Se conservan las NC/ND SOLO_GESTION que SÍ tienen producto identificado
    # — esas son los hallazgos reales (NC para producto Syngenta que el
    # distribuidor no le reportó a Syngenta).
    conciliacion = [
        r for r in conciliacion
        if not (
            r["estado"] == "SOLO_GESTION"
            and str(r.get("tipo_comprobante", "")).upper() in ("NC", "ND")
            and not str(r.get("producto") or "").strip()
        )
    ]

    conciliacion.sort(key=lambda x: x.get("fecha", ""))
    return conciliacion
